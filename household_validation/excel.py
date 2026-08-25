from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from household_validation.identity import get_household_form_number


YES_NO_FORMULA = '"YES,NO"'
PRIMARY_WORKER_FORMULA = '"YES,NO"'

LOCATION_COLUMN_TYPES = {
    "District": "R",
    "TA": "D",
    "GVH": "W",
    "Village": "V",
}

MICRO_CATCHMENT_COLUMN = "Micro-Catchment"
HOTSPOT_COLUMN = "Hotspot"


EXCEL_COLUMNS = [
    "batch_id",
    "group_uuid",
    "member_uuid",
    "row_type",
    "District",
    MICRO_CATCHMENT_COLUMN,
    "TA",
    "GVH",
    HOTSPOT_COLUMN,
    "Village",
    "form_number",
    "member_name",
    "relationship",
    "member_dob",
    "national_id",
    "primary_worker",
    "verified",
    "member_gender",
    "member_age",
    "marital_status",
    "disability",
    "fit_for_work",
    "head",
    "pmt_score",
    "household_wealth_quintile",
    "validation_date",
    "project",
    "project_id",
    "validation_notes",
]

PROJECT_OPTIONS_SHEET = "Project Options"
PROJECT_OPTIONS_HEADERS = ["project_id", "project", "project_label"]

EDITABLE_COLUMNS = {
    "primary_worker",
    "verified",
    "validation_date",
    "project",
    "validation_notes",
}

TEXT_COLUMNS = {
    "form_number",
    "national_id",
}


class ExcelValidationListExporter:
    def __init__(self, selection_result, batch_id, projects=None):
        self.selection_result = selection_result
        self.batch_id = batch_id
        self.projects = projects or []
        self._micro_catchment_cache = {}
        self._hotspot_cache = {}

    def export_workbook(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Validation List"

        self._write_header(worksheet)
        self._write_rows(worksheet)
        project_options_worksheet = self._write_project_options(workbook)
        self._apply_validation(worksheet)
        self._apply_protection(worksheet)
        self._autosize_columns(worksheet)
        self._autosize_columns(project_options_worksheet)

        project_id_column = EXCEL_COLUMNS.index("project_id") + 1
        worksheet.column_dimensions[worksheet.cell(1, project_id_column).column_letter].hidden = True
        return workbook

    def export_bytes(self):
        output = BytesIO()
        self.export_workbook().save(output)
        output.seek(0)
        return output.getvalue()

    def _write_header(self, worksheet):
        header_fill = PatternFill("solid", fgColor="D9EAD3")
        for column_number, title in enumerate(EXCEL_COLUMNS, start=1):
            cell = worksheet.cell(row=1, column=column_number, value=title)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.protection = Protection(locked=True)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

    def _write_rows(self, worksheet):
        for row_number, selected_member in enumerate(self.selection_result.member_rows, start=2):
            values = self._build_row(selected_member)
            for column_number, title in enumerate(EXCEL_COLUMNS, start=1):
                value = values.get(title)
                if title in TEXT_COLUMNS and value is not None:
                    value = str(value)
                cell = worksheet.cell(
                    row=row_number,
                    column=column_number,
                    value=value,
                )
                if title in TEXT_COLUMNS:
                    cell.number_format = "@"
                cell.protection = Protection(locked=title not in EDITABLE_COLUMNS)

    def _write_project_options(self, workbook):
        worksheet = workbook.create_sheet(PROJECT_OPTIONS_SHEET)
        for column_number, title in enumerate(PROJECT_OPTIONS_HEADERS, start=1):
            worksheet.cell(row=1, column=column_number, value=title)
        project_labels = self._project_labels()
        for row_number, project in enumerate(self.projects, start=2):
            worksheet.cell(row=row_number, column=1, value=self._project_id(project))
            worksheet.cell(row=row_number, column=2, value=self._project_name(project))
            worksheet.cell(row=row_number, column=3, value=project_labels[id(project)])
        worksheet.sheet_state = "hidden"
        return worksheet

    def _build_row(self, selected_member):
        household = selected_member.household
        member = selected_member.member
        group = household.source
        group_individual = member.source
        individual = getattr(group_individual, "individual", None)
        location = getattr(group, "location", None)

        return {
            "batch_id": str(self.batch_id),
            "row_type": selected_member.row_type,
            **{
                column: self._location_name(location, location_type)
                for column, location_type in LOCATION_COLUMN_TYPES.items()
            },
            MICRO_CATCHMENT_COLUMN: self._micro_catchment_name(location),
            HOTSPOT_COLUMN: self._hotspot_name(location),
            "form_number": get_household_form_number(group, individual),
            "group_uuid": str(household.id),
            "member_uuid": str(member.id),
            "member_name": self._member_name(individual),
            "national_id": self._national_id(individual),
            "member_gender": member.gender,
            "member_dob": member.dob,
            "member_age": member.age,
            "marital_status": self._marital_status(individual),
            "disability": self._disability(individual),
            "fit_for_work": "YES" if member.fit_for_work else "NO",
            "relationship": self._relationship(member.role),
            "head": "YES" if self._is_head(member) else "NO",
            "pmt_score": household.pmt_score,
            "household_wealth_quintile": household.wealth_quintile,
            "primary_worker": self._primary_worker(group_individual),
            "verified": None,
            "validation_date": None,
            "project": None,
            "project_id": None,
            "validation_notes": None,
        }

    def _apply_validation(self, worksheet):
        max_row = max(worksheet.max_row, 2)
        primary_worker_col = self._column_letter("primary_worker")
        verified_col = self._column_letter("verified")
        project_col = self._column_letter("project")

        primary_worker_validation = DataValidation(
            type="list",
            formula1=PRIMARY_WORKER_FORMULA,
            allow_blank=True,
        )
        verified_validation = DataValidation(
            type="list",
            formula1=YES_NO_FORMULA,
            allow_blank=True,
        )

        worksheet.add_data_validation(primary_worker_validation)
        worksheet.add_data_validation(verified_validation)
        primary_worker_validation.add(
            f"{primary_worker_col}2:{primary_worker_col}{max_row}"
        )
        verified_validation.add(f"{verified_col}2:{verified_col}{max_row}")

        project_count = len([project for project in self.projects if self._project_name(project)])
        if project_count:
            project_formula = f"'{PROJECT_OPTIONS_SHEET}'!$C$2:$C${project_count + 1}"
            project_validation = DataValidation(
                type="list",
                formula1=project_formula,
                allow_blank=True,
            )
            worksheet.add_data_validation(project_validation)
            project_validation.add(f"{project_col}2:{project_col}{max_row}")

    def _apply_protection(self, worksheet):
        worksheet.protection.sheet = True
        worksheet.protection.enable()

    def _autosize_columns(self, worksheet):
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = cell.value
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)

    def _column_letter(self, column_name):
        return get_column_letter(EXCEL_COLUMNS.index(column_name) + 1)

    def _location_name(self, location, location_type):
        current = location
        while current is not None:
            if getattr(current, "type", None) == location_type:
                return getattr(current, "name", None) or getattr(current, "code", None)
            current = getattr(current, "parent", None)
        return None

    def _micro_catchment_name(self, location):
        """Micro-catchment for the row's location.

        A micro-catchment isn't a level in the Region/TA/GVH/Village chain — it's
        a separate grouping of specific GVHs (and TAs) defined in ``location.MicroCatchment``.
        Prefer a GVH-level link (more specific) and fall back to the TA-level link.
        """
        gvh_location = self._location_ancestor(location, "W")
        if gvh_location is not None:
            name = self._micro_catchment_link_name(gvh_location, "micro_catchments_gvh")
            if name:
                return name

        ta_location = self._location_ancestor(location, "D")
        if ta_location is not None:
            return self._micro_catchment_link_name(ta_location, "micro_catchments_ta")

        return None

    def _location_ancestor(self, location, location_type):
        current = location
        while current is not None:
            if getattr(current, "type", None) == location_type:
                return current
            current = getattr(current, "parent", None)
        return None

    def _micro_catchment_link_name(self, location, related_name):
        related_manager = getattr(location, related_name, None)
        location_id = getattr(location, "id", None)
        if related_manager is None or location_id is None:
            return None
        cache_key = (related_name, location_id)
        if cache_key not in self._micro_catchment_cache:
            link = related_manager.filter(
                validity_to__isnull=True,
                micro_catchment__validity_to__isnull=True,
            ).select_related("micro_catchment").first()
            micro_catchment = link.micro_catchment if link else None
            name = None
            if micro_catchment is not None:
                name = micro_catchment.name or micro_catchment.code
            self._micro_catchment_cache[cache_key] = name
        return self._micro_catchment_cache[cache_key]

    def _hotspot_name(self, location):
        """Hotspot for the row's location.

        A hotspot links to specific villages (``location.HotspotVillage``), so unlike
        the micro-catchment lookup there's no ancestor tier to fall back through —
        just resolve the village-level ancestor's hotspot link, if any.
        """
        village_location = self._location_ancestor(location, "V")
        if village_location is None:
            return None

        related_manager = getattr(village_location, "hotspot_links", None)
        location_id = getattr(village_location, "id", None)
        if related_manager is None or location_id is None:
            return None
        if location_id not in self._hotspot_cache:
            link = related_manager.filter(
                validity_to__isnull=True,
                hotspot__validity_to__isnull=True,
            ).select_related("hotspot").first()
            hotspot = link.hotspot if link else None
            name = None
            if hotspot is not None:
                name = hotspot.name or hotspot.code
            self._hotspot_cache[location_id] = name
        return self._hotspot_cache[location_id]

    def _member_name(self, individual):
        if not individual:
            return None
        first_name = getattr(individual, "first_name", "") or ""
        last_name = getattr(individual, "last_name", "") or ""
        return f"{first_name} {last_name}".strip()

    def _national_id(self, individual):
        if not individual:
            return None
        return (getattr(individual, "json_ext", None) or {}).get("national_id")

    def _marital_status(self, individual):
        if not individual:
            return None
        return (getattr(individual, "json_ext", None) or {}).get("marital_status")

    def _disability(self, individual):
        if not individual:
            return None
        return (getattr(individual, "json_ext", None) or {}).get("disability")

    def _primary_worker(self, group_individual):
        primary_worker = (
            getattr(group_individual, "json_ext", None) or {}
        ).get("primary_worker")
        if primary_worker is True:
            return "YES"
        if primary_worker is False:
            return "NO"
        recipient_type = str(getattr(group_individual, "recipient_type", "") or "").upper()
        if recipient_type == "PRIMARY":
            return "YES"
        return None

    def _relationship(self, role):
        if role is None:
            return None
        return str(role).strip() or None

    def _is_head(self, member):
        return str(member.role or "").upper() == "HEAD"

    def _project_name(self, project):
        return getattr(project, "name", None)

    def _project_id(self, project):
        return str(getattr(project, "id", "") or getattr(project, "uuid", "") or "")

    def _project_labels(self):
        name_counts = {}
        for project in self.projects:
            project_name = self._project_name(project)
            if project_name:
                name_counts[project_name] = name_counts.get(project_name, 0) + 1

        labels = {}
        for project in self.projects:
            project_name = self._project_name(project)
            project_id = self._project_id(project)
            if project_name and name_counts.get(project_name, 0) > 1 and project_id:
                labels[id(project)] = f"{project_name} ({project_id})"
            else:
                labels[id(project)] = project_name
        return labels
