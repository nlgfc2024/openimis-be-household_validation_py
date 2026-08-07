from dataclasses import dataclass, field
import csv
from datetime import date, datetime
from io import BytesIO
from io import StringIO

from openpyxl import load_workbook

from household_validation.excel import (
    EXCEL_COLUMNS,
    PROJECT_OPTIONS_SHEET,
    PROJECT_OPTIONS_HEADERS,
)


VALIDATION_LIST_SHEET = "Validation List"
PROJECT_SELECTION_TYPE_INTENT = "INTENT"
VALIDATION_STATUS_VERIFIED = "VERIFIED"
VALIDATION_STATUS_NOT_VERIFIED = "NOT_VERIFIED"

EDITABLE_UPLOAD_COLUMNS = {
    "primary_worker",
    "verified",
    "validation_date",
    "project",
    "validation_notes",
}
STRUCTURAL_UPLOAD_COLUMNS = tuple(
    column for column in EXCEL_COLUMNS if column not in EDITABLE_UPLOAD_COLUMNS
)

YES_VALUES = {"YES", "Y", "TRUE", "1"}
NO_VALUES = {"NO", "N", "FALSE", "0"}


@dataclass(frozen=True)
class UploadedValidationRow:
    row_number: int
    values: dict
    verified: bool | None
    participant: bool
    validation_date: date | None
    project_name: str | None
    project_id: str | None
    notes: str | None


@dataclass(frozen=True)
class WorkbookParseResult:
    rows: list[UploadedValidationRow] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    project_options: dict = field(default_factory=dict)

    @property
    def rows_read(self):
        return len(self.rows)


def parse_validation_workbook(file_or_bytes):
    workbook = load_workbook(_to_bytes_io(file_or_bytes), data_only=True)
    errors = []
    if VALIDATION_LIST_SHEET not in workbook.sheetnames:
        return WorkbookParseResult(errors=[f"Missing worksheet: {VALIDATION_LIST_SHEET}"])

    worksheet = workbook[VALIDATION_LIST_SHEET]
    headers = _read_headers(worksheet)
    missing_columns = [column for column in EXCEL_COLUMNS if column not in headers]
    if missing_columns:
        return WorkbookParseResult(
            errors=[f"Missing required columns: {', '.join(missing_columns)}"]
        )

    project_options = _read_project_options(workbook)
    rows = []
    for row_number in range(2, worksheet.max_row + 1):
        values = {
            column: _cell_value(worksheet.cell(row=row_number, column=headers[column]))
            for column in EXCEL_COLUMNS
        }
        if _is_blank_row(values):
            continue
        row_errors = _validate_structural_values(row_number, values)
        verified = _parse_yes_no(values.get("verified"))
        participant = _parse_yes_no(values.get("primary_worker")) is True
        validation_date = _parse_date(values.get("validation_date"))
        project_label = _clean(values.get("project"))
        project_name = _resolve_project_name(project_label, project_options)
        project_id = _resolve_project_id(
            project_label=project_label,
            workbook_project_id=_clean(values.get("project_id")),
            project_options=project_options,
        )
        if values.get("project_id") and not project_label:
            row_errors.append(f"Row {row_number}: project_id cannot be set without project")
        if values.get("verified") not in (None, "") and verified is None:
            row_errors.append(f"Row {row_number}: verified must be YES or NO")
        if values.get("primary_worker") not in (None, "") and _parse_yes_no(
            values.get("primary_worker")
        ) is None:
            row_errors.append(f"Row {row_number}: primary_worker must be YES or NO")
        if values.get("validation_date") and validation_date is None:
            row_errors.append(f"Row {row_number}: validation_date is invalid")
        if project_label and not project_id:
            row_errors.append(f"Row {row_number}: project is not in the project options")
        if row_errors:
            errors.extend(row_errors)
            continue
        rows.append(
            UploadedValidationRow(
                row_number=row_number,
                values=values,
                verified=verified,
                participant=participant,
                validation_date=validation_date,
                project_name=project_name,
                project_id=project_id,
                notes=_clean(values.get("validation_notes")),
            )
        )
    return WorkbookParseResult(rows=rows, errors=errors, project_options=project_options)


def _to_bytes_io(file_or_bytes):
    if isinstance(file_or_bytes, bytes):
        return BytesIO(file_or_bytes)
    if hasattr(file_or_bytes, "read"):
        content = file_or_bytes.read()
        if hasattr(file_or_bytes, "seek"):
            file_or_bytes.seek(0)
        return BytesIO(content)
    return file_or_bytes


def _read_headers(worksheet):
    headers = {}
    for column_number in range(1, worksheet.max_column + 1):
        value = _clean(worksheet.cell(row=1, column=column_number).value)
        if value:
            headers[value] = column_number
    return headers


def _read_project_options(workbook):
    if PROJECT_OPTIONS_SHEET not in workbook.sheetnames:
        return {}
    worksheet = workbook[PROJECT_OPTIONS_SHEET]
    headers = _read_headers(worksheet)
    if not all(header in headers for header in PROJECT_OPTIONS_HEADERS[:2]):
        return {}
    options = {}
    for row_number in range(2, worksheet.max_row + 1):
        project_id = _clean(worksheet.cell(row=row_number, column=headers["project_id"]).value)
        project_name = _clean(worksheet.cell(row=row_number, column=headers["project"]).value)
        project_label = project_name
        if "project_label" in headers:
            project_label = _clean(worksheet.cell(row=row_number, column=headers["project_label"]).value)
        if project_id and project_name:
            options[project_label or project_name] = {
                "id": project_id,
                "name": project_name,
            }
    return options


def _validate_structural_values(row_number, values):
    errors = []
    for column in ("batch_id", "group_uuid", "member_uuid"):
        if not _clean(values.get(column)):
            errors.append(f"Row {row_number}: {column} is required")
    return errors


def _is_blank_row(values):
    return all(value in (None, "") for value in values.values())


def _cell_value(cell):
    value = cell.value
    if isinstance(value, datetime):
        return value.date()
    return value


def _clean(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def _parse_yes_no(value):
    value = _clean(value)
    if value is None:
        return None
    normalized = value.upper()
    if normalized in YES_VALUES:
        return True
    if normalized in NO_VALUES:
        return False
    return None


def _parse_date(value):
    if isinstance(value, date):
        return value
    value = _clean(value)
    if not value:
        return None
    try:
        return date.fromisoformat(value[:10])
    except ValueError:
        return None


def _resolve_project_name(project_label, project_options):
    if not project_label:
        return None
    option = project_options.get(project_label)
    if isinstance(option, dict):
        return option.get("name")
    if option:
        return project_label
    return None


def _resolve_project_id(project_label, workbook_project_id, project_options):
    if not project_label:
        return workbook_project_id
    option = project_options.get(project_label)
    project_id = option.get("id") if isinstance(option, dict) else option
    if workbook_project_id and project_id and workbook_project_id != project_id:
        return None
    return project_id or workbook_project_id


def build_validation_json_ext(uploaded_row, project, upload_date, uploaded_at, user_id):
    validation_date = uploaded_row.validation_date or upload_date
    validation_status = (
        VALIDATION_STATUS_VERIFIED
        if uploaded_row.verified is True
        else VALIDATION_STATUS_NOT_VERIFIED
    )
    return {
        "validation_status": validation_status,
        "last_verified_date": validation_date.isoformat(),
        "validation_project_id": str(project.id) if project else uploaded_row.project_id,
        "validation_project_name": project.name if project else uploaded_row.project_name,
        "validation_project_selection_type": PROJECT_SELECTION_TYPE_INTENT,
        "validation_uploaded_at": uploaded_at.isoformat(),
        "validation_uploaded_by_id": str(user_id) if user_id else None,
        "validation_notes": uploaded_row.notes,
    }


def member_structural_errors(uploaded_row, group_individual, group=None):
    errors = []
    row_number = uploaded_row.row_number
    individual = getattr(group_individual, "individual", None)
    individual_json_ext = getattr(individual, "json_ext", None) or {}
    group = group or getattr(group_individual, "group", None)

    expected = {
        "form_number": _form_number(group, individual),
        "member_name": _member_name(individual),
        "national_id": individual_json_ext.get("national_id"),
        "member_gender": individual_json_ext.get("gender"),
        "member_dob": _date_value(getattr(individual, "dob", None)),
        "member_age": _age(getattr(individual, "dob", None)),
        "fit_for_work": "YES" if _truthy(individual_json_ext.get("fit_for_work")) else "NO",
        "relationship": _relationship(getattr(group_individual, "role", None)),
        "head": "YES" if str(getattr(group_individual, "role", "")).upper() == "HEAD" else "NO",
        "household_wealth_quintile": _household_wealth_quintile(group),
        "current_recipient_type": getattr(group_individual, "recipient_type", None),
    }
    strict_columns = {
        "form_number",
        "national_id",
        "relationship",
        "household_wealth_quintile",
    }
    for column, expected_value in expected.items():
        uploaded_value = uploaded_row.values.get(column)
        if expected_value in (None, ""):
            continue
        if column not in strict_columns and uploaded_value in (None, ""):
            continue
        if _normalize(uploaded_value) != _normalize(expected_value):
            errors.append(f"Row {row_number}: {column} does not match the household member")
    return errors


def build_validation_error_report_csv(batch_id, rows):
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "batch_id",
            "row_number",
            "status",
            "form_number",
            "group_uuid",
            "member_uuid",
            "error_message",
        ]
    )
    for row in rows:
        raw_row = row.raw_row or {}
        writer.writerow(
            [
                str(batch_id),
                row.row_number,
                row.status,
                raw_row.get("form_number"),
                raw_row.get("group_uuid"),
                raw_row.get("member_uuid"),
                row.error_message,
            ]
        )
    return output.getvalue()


def _member_name(individual):
    if individual is None:
        return None
    return f"{getattr(individual, 'first_name', '') or ''} {getattr(individual, 'last_name', '') or ''}".strip()


def _relationship(role):
    if role is None:
        return None
    return str(role).strip() or None


def _form_number(group, individual):
    group_json_ext = getattr(group, "json_ext", None) or {}
    individual_json_ext = getattr(individual, "json_ext", None) or {}
    return group_json_ext.get("form_number") or individual_json_ext.get("form_number")


def _household_wealth_quintile(group):
    if group is None:
        return None
    group_json_ext = getattr(group, "json_ext", None) or {}
    if group_json_ext.get("household_wealth_quintile") is not None:
        return group_json_ext.get("household_wealth_quintile")

    group_individuals = getattr(group, "groupindividuals", None)
    if group_individuals is None:
        return None
    members = group_individuals.all()
    for member in members:
        individual = getattr(member, "individual", None)
        individual_json_ext = getattr(individual, "json_ext", None) or {}
        if individual_json_ext.get("household_wealth_quintile") is not None:
            return individual_json_ext.get("household_wealth_quintile")
    return None


def _date_value(value):
    if isinstance(value, date):
        return value.isoformat()
    return value


def _age(dob):
    if not isinstance(dob, date):
        return None
    today = date.today()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))


def _truthy(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y"}
    return bool(value)


def _normalize(value):
    if value is None:
        return None
    if isinstance(value, date):
        value = value.isoformat()
    value = str(value).strip().casefold()
    return value or None
