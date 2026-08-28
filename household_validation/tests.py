import base64
from contextlib import nullcontext
from datetime import date, datetime, timezone as datetime_timezone
from io import BytesIO
from types import SimpleNamespace
from unittest import TestCase
from unittest.mock import MagicMock, call, patch
from uuid import uuid4

from django.db.models import Q
from django.test import TestCase as DjangoTestCase
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from core.models import User
from individual.models import Group, GroupIndividual, Individual
from location.models import Hotspot, HotspotVillage, Location, MicroCatchment

from household_validation.apps import (
    DEFAULT_CONFIG,
    DISTRICT_VALIDATION_ROLE_RIGHTS,
    DISTRICT_VALIDATION_ROLES,
    GROUP_RIGHTS,
    HOUSEHOLD_VALIDATION_RIGHTS,
    ROLE_DISTRICT_ADMINISTRATOR,
    ROLE_DISTRICT_PROGRAM_MANAGER,
    ROLE_DISTRICT_USER,
    RIGHT_GROUP_SEARCH,
    RIGHT_GROUP_UPDATE,
    RIGHT_HOUSEHOLD_VALIDATION_ERROR_REPORT,
    RIGHT_HOUSEHOLD_VALIDATION_HISTORY,
    RIGHT_HOUSEHOLD_VALIDATION_QUERY_EXPORT,
    RIGHT_HOUSEHOLD_VALIDATION_UPLOAD,
    HouseholdValidationConfig,
)
from household_validation.excel import (
    EXCEL_COLUMNS,
    HOTSPOT_COLUMN,
    MICRO_CATCHMENT_COLUMN,
    PROJECT_OPTIONS_SHEET,
    ExcelValidationListExporter,
    build_rejected_households_workbook_bytes,
)
from household_validation.gql_permissions import (
    HouseholdValidationPermissionError,
    require_permissions,
)
from household_validation.gql_mutations import GenerateHouseholdValidationListMutation
from household_validation.models import HouseholdValidationBatchRow
from household_validation.project_lookup import (
    ACTIVE_PROJECT_STATUSES,
    ProjectOption,
    project_option_from_project,
)
from household_validation.selection import (
    CATEGORY_FEMALE_HEADED,
    CATEGORY_OTHER,
    CATEGORY_YOUTH,
    ROW_TYPE_MAIN,
    ROW_TYPE_RESERVE,
    EligibleHousehold,
    EligibleMember,
    SelectedHousehold,
    SelectionResult,
    is_truthy,
    select_households,
)
from household_validation.services import (
    EligibleHouseholdSelectionService,
    HouseholdValidationUploadService,
    _json_safe,
    _local_date,
)
from household_validation.schema import Query
from household_validation.upload import (
    OPTIONAL_UPLOAD_COLUMNS,
    PROJECT_SELECTION_TYPE_INTENT,
    UploadedValidationRow,
    VALIDATION_STATUS_NOT_VERIFIED,
    VALIDATION_LIST_SHEET,
    build_validation_error_report_csv,
    build_validation_json_ext,
    member_structural_errors,
    parse_validation_workbook,
)
from household_validation.wealth import get_household_pmt_score, get_household_wealth_quintile


class HouseholdValidationConfigTest(TestCase):
    def test_default_config_exposes_permission_lists(self):
        self.assertEqual(
            DEFAULT_CONFIG["gql_query_household_validation_rule_perms"],
            [str(RIGHT_HOUSEHOLD_VALIDATION_QUERY_EXPORT)],
        )
        self.assertEqual(
            DEFAULT_CONFIG["gql_mutation_generate_household_validation_list_perms"],
            [str(RIGHT_HOUSEHOLD_VALIDATION_QUERY_EXPORT)],
        )
        self.assertEqual(
            DEFAULT_CONFIG["gql_mutation_upload_household_validation_list_perms"],
            [str(RIGHT_HOUSEHOLD_VALIDATION_UPLOAD)],
        )
        self.assertEqual(
            DEFAULT_CONFIG["gql_query_household_validation_history_perms"],
            [str(RIGHT_HOUSEHOLD_VALIDATION_HISTORY)],
        )
        self.assertEqual(
            DEFAULT_CONFIG["gql_query_household_validation_error_report_perms"],
            [str(RIGHT_HOUSEHOLD_VALIDATION_ERROR_REPORT)],
        )
        self.assertEqual(
            DEFAULT_CONFIG["group_search_perms"],
            [str(RIGHT_GROUP_SEARCH)],
        )
        self.assertEqual(
            DEFAULT_CONFIG["group_update_perms"],
            [str(RIGHT_GROUP_UPDATE)],
        )

    def test_default_config_exposes_selection_percentages(self):
        self.assertEqual(DEFAULT_CONFIG["female_headed_percentage"], 40)
        self.assertEqual(DEFAULT_CONFIG["youth_percentage"], 40)
        self.assertEqual(DEFAULT_CONFIG["reserve_percentage"], 20)

    def test_right_sets_match_default_config_scope(self):
        self.assertEqual(
            HOUSEHOLD_VALIDATION_RIGHTS,
            [
                RIGHT_HOUSEHOLD_VALIDATION_QUERY_EXPORT,
                RIGHT_HOUSEHOLD_VALIDATION_UPLOAD,
                RIGHT_HOUSEHOLD_VALIDATION_HISTORY,
                RIGHT_HOUSEHOLD_VALIDATION_ERROR_REPORT,
            ],
        )
        self.assertIn(RIGHT_GROUP_SEARCH, GROUP_RIGHTS)
        self.assertIn(RIGHT_GROUP_UPDATE, GROUP_RIGHTS)

    def test_district_validation_roles_have_validation_and_group_update_rights(self):
        self.assertEqual(
            DISTRICT_VALIDATION_ROLES,
            [
                ROLE_DISTRICT_ADMINISTRATOR,
                ROLE_DISTRICT_PROGRAM_MANAGER,
                ROLE_DISTRICT_USER,
            ],
        )
        for role_name in DISTRICT_VALIDATION_ROLES:
            role_rights = DISTRICT_VALIDATION_ROLE_RIGHTS[role_name]
            for right in HOUSEHOLD_VALIDATION_RIGHTS:
                self.assertIn(right, role_rights)
            self.assertIn(RIGHT_GROUP_SEARCH, role_rights)
            self.assertIn(RIGHT_GROUP_UPDATE, role_rights)


class RejectedBatchRowsQueryTest(TestCase):
    @patch("household_validation.schema.HouseholdValidationBatchRow.objects.filter")
    def test_rejected_rows_use_upload_permission_without_changing_history_query(
        self,
        filter_mock,
    ):
        user = MagicMock()
        info = MagicMock()
        info.context.user = user
        rejected_row = MagicMock()
        rejected_row.json_ext = {"error_code": "MULTIPLE_PRIMARY_WORKERS"}
        rejected_row.raw_row = {
            "form_number": "FORM-001",
            "group_uuid": "group-1",
        }
        rejected_row.group_id = "group-1"
        rejected_row.row_number = 2
        rejected_row.error_message = (
            "household has more than one primary worker"
        )
        upload_attempt_id = "01a03ab4-c5ac-7b78-a485-f321e7d092f8"
        filter_mock.return_value.order_by.return_value = [rejected_row]

        with patch.object(Query, "_check_permissions") as check_permissions_mock:
            result = Query.resolve_household_validation_rejected_batch_rows(
                None,
                info,
                batch_id="01a03aae-4896-7f00-9357-78f69fb8e6ca",
                upload_attempt_id=upload_attempt_id,
            )

        check_permissions_mock.assert_called_once_with(
            user,
            HouseholdValidationConfig.gql_mutation_upload_household_validation_list_perms,
        )
        filter_mock.assert_called_once_with(
            batch_id="01a03aae-4896-7f00-9357-78f69fb8e6ca",
            upload_attempt_id=upload_attempt_id,
            status=HouseholdValidationBatchRow.Status.ERROR,
            is_deleted=False,
        )
        filter_mock.return_value.order_by.assert_called_once_with("row_number")
        self.assertEqual(result.rows, [rejected_row])
        self.assertEqual(result.count, 1)
        self.assertEqual(
            result.file_name,
            (
                "rejected_households_01a03aae-4896-7f00-9357-78f69fb8e6ca_"
                "01a03ab4-c5ac-7b78-a485-f321e7d092f8.xlsx"
            ),
        )
        workbook = load_workbook(BytesIO(base64.b64decode(result.file_base64)))
        self.assertEqual(workbook.active.title, "Rejected Households")
        self.assertEqual(workbook.active.max_row, 2)


class RejectedHouseholdsWorkbookTest(TestCase):
    def test_contains_only_multiple_primary_worker_rejections(self):
        rejection_message = (
            "household has more than one primary worker"
        )
        rows = [
            SimpleNamespace(
                row_number=4,
                group_id="group-1",
                raw_row={"form_number": "FORM-001", "member_name": "Member One"},
                json_ext={"error_code": "MULTIPLE_PRIMARY_WORKERS"},
                error_message=rejection_message,
            ),
            SimpleNamespace(
                row_number=5,
                group_id="group-1",
                raw_row={"form_number": "FORM-001", "member_name": "Member Two"},
                json_ext={"error_code": "MULTIPLE_PRIMARY_WORKERS"},
                error_message=rejection_message,
            ),
            SimpleNamespace(
                row_number=6,
                group_id="group-2",
                raw_row={"form_number": "FORM-002", "member_name": "Other Error"},
                json_ext={"error_code": "MEMBER_NOT_FOUND"},
                error_message="Member was not found",
            ),
        ]

        report, rejected_row_count = build_rejected_households_workbook_bytes(rows)
        workbook = load_workbook(BytesIO(report))
        worksheet = workbook["Rejected Households"]
        values = list(worksheet.values)

        self.assertEqual(rejected_row_count, 1)
        self.assertEqual(worksheet.max_row, 2)
        self.assertIn("rejection_reason", values[0])
        self.assertEqual(values[1][0], "FORM-001")
        self.assertEqual(values[1][1], "group-1")
        self.assertEqual(values[1][2], "4, 5")
        self.assertNotIn("FORM-002", str(values))


def _dob_for_age(age):
    today = date.today()
    return date(today.year - age, today.month, today.day)


def _member(member_id, gender="Male", age=40, fit_for_work=True, role=None):
    return EligibleMember(
        id=member_id,
        gender=gender,
        dob=_dob_for_age(age),
        fit_for_work=fit_for_work,
        role=role,
    )


def _household(
    household_id,
    wealth_quintile,
    head_gender="Male",
    eligible_member_age=40,
    last_verified_date=None,
    eligible_members=None,
    village_id=None,
    village_code=None,
    village_name=None,
):
    head = _member(
        f"{household_id}-head",
        gender=head_gender,
        age=50,
        role="HEAD",
    )
    if eligible_members is None:
        eligible_members = [
            _member(
                f"{household_id}-worker",
                gender="Male",
                age=eligible_member_age,
            )
        ]
    return EligibleHousehold(
        id=household_id,
        code=str(household_id),
        wealth_quintile=wealth_quintile,
        last_verified_date=last_verified_date,
        head=head,
        eligible_members=eligible_members,
        village_id=village_id,
        village_code=village_code,
        village_name=village_name,
    )


class HouseholdSelectionTest(TestCase):
    def test_select_households_allocates_target_proportionally_by_village(self):
        households = (
            [
                _household(f"a-{index}", "Poorest", village_code="A", village_name="Village A")
                for index in range(30)
            ]
            + [
                _household(f"b-{index}", "Poorest", village_code="B", village_name="Village B")
                for index in range(15)
            ]
            + [
                _household(f"c-{index}", "Poorest", village_code="C", village_name="Village C")
                for index in range(10)
            ]
        )

        result, summary = select_households(
            households,
            target_count=20,
            allocate_by_village=True,
        )

        selected_by_village = {
            code: len([
                row
                for row in result.main
                if row.household.village_code == code
            ])
            for code in ("A", "B", "C")
        }
        self.assertEqual(selected_by_village, {"A": 11, "B": 5, "C": 4})
        self.assertEqual(summary["selected_households"], 20)
        self.assertEqual(
            {
                row["village_code"]: row["allocated_households"]
                for row in summary["village_breakdown"]
            },
            {"A": 11, "B": 5, "C": 4},
        )

    def test_village_allocation_guarantees_one_household_when_target_allows(self):
        households = (
            [
                _household(f"a-{index}", "Poorest", village_code="A")
                for index in range(98)
            ]
            + [_household("b-1", "Poorest", village_code="B")]
            + [_household("c-1", "Poorest", village_code="C")]
        )

        result, _ = select_households(
            households,
            target_count=3,
            allocate_by_village=True,
        )

        self.assertEqual(
            {row.household.village_code for row in result.main},
            {"A", "B", "C"},
        )

    def test_village_reserve_is_proportional_and_excludes_main_households(self):
        households = [
            _household(
                f"{code}-{index}",
                "Poorest",
                village_code=code,
            )
            for code in ("A", "B")
            for index in range(10)
        ]

        with patch.object(HouseholdValidationConfig, "reserve_percentage", 20):
            result, summary = select_households(
                households,
                target_count=10,
                allocate_by_village=True,
            )

        self.assertEqual(
            {
                code: len([
                    row
                    for row in result.main
                    if row.household.village_code == code
                ])
                for code in ("A", "B")
            },
            {"A": 5, "B": 5},
        )
        self.assertEqual(
            {
                code: len([
                    row
                    for row in result.reserve
                    if row.household.village_code == code
                ])
                for code in ("A", "B")
            },
            {"A": 1, "B": 1},
        )
        self.assertFalse(
            {row.household.id for row in result.main}
            & {row.household.id for row in result.reserve}
        )
        self.assertEqual(summary["reserve_households"], 2)

    def test_select_households_applies_quota_categories(self):
        households = [
            _household("female", "Poorest", head_gender="Female"),
            _household("youth", "Poorest", eligible_member_age=25),
            _household("other", "Poorest"),
        ]

        result, _ = select_households(households, target_count=3)

        self.assertEqual([row.row_type for row in result.main], [ROW_TYPE_MAIN] * 3)
        self.assertEqual(
            [row.category for row in result.main],
            [CATEGORY_FEMALE_HEADED, CATEGORY_YOUTH, CATEGORY_OTHER],
        )

    def test_select_households_prioritizes_poorest_within_category(self):
        households = [
            _household("female-richer", "Richer", head_gender="Female"),
            _household("female-poorest", "Poorest", head_gender="Female"),
            _household("youth", "Poorest", eligible_member_age=20),
        ]

        result, _ = select_households(households, target_count=1)

        self.assertEqual(result.main[0].household.id, "female-poorest")

    def test_select_households_defaults_target_to_all_eligible_and_caps_reserve(self):
        households = [
            _household("female", "Poorest", head_gender="Female"),
            _household("youth", "Poorer", eligible_member_age=20),
        ]

        result, _ = select_households(households)

        self.assertEqual(len(result.main), 2)
        self.assertEqual(result.reserve, [])

    def test_select_households_adds_reserve_from_remaining_households(self):
        households = [
            _household("female", "Poorest", head_gender="Female"),
            _household("youth", "Poorer", eligible_member_age=20),
            _household("other", "Middle"),
        ]

        with patch.object(HouseholdValidationConfig, "reserve_percentage", 50):
            result, summary = select_households(households, target_count=2)

        self.assertEqual(len(result.main), 2)
        self.assertEqual(len(result.reserve), 1)
        self.assertEqual(result.reserve[0].row_type, ROW_TYPE_RESERVE)
        self.assertEqual(result.reserve[0].household.id, "other")
        self.assertEqual(summary["selected_households"], 2)
        self.assertEqual(summary["selected_individuals"], 2)
        self.assertEqual(summary["reserve_households"], 1)

    def test_select_households_adds_default_twenty_percent_reserve(self):
        households = [
            _household(index, "Poorest")
            for index in range(1, 21)
        ]

        result, _ = select_households(households, target_count=10)

        self.assertEqual(len(result.main), 10)
        self.assertEqual(len(result.reserve), 2)
        self.assertEqual(result.reserve[0].row_type, ROW_TYPE_RESERVE)

    def test_reserve_capacity_uses_actual_main_selection_count(self):
        households = [
            _household(f"household-{index}", "Poorest")
            for index in range(3)
        ]
        selected_main = [
            SelectedHousehold(
                household=household,
                category=CATEGORY_OTHER,
                row_type=ROW_TYPE_MAIN,
            )
            for household in households[:2]
        ]
        category_counts = {
            CATEGORY_FEMALE_HEADED: 0,
            CATEGORY_YOUTH: 0,
            CATEGORY_OTHER: 2,
        }
        household_categories = {
            household.id: CATEGORY_OTHER for household in households
        }

        with (
            patch(
                "household_validation.selection._select_main_households",
                return_value=(
                    selected_main,
                    {household.id for household in households[:2]},
                    category_counts,
                    sum(
                        len(household.eligible_members)
                        for household in households[:2]
                    ),
                    household_categories,
                ),
            ),
            patch.object(HouseholdValidationConfig, "reserve_percentage", 100),
        ):
            result, summary = select_households(households, target_count=3)

        self.assertEqual(
            [row.household.id for row in result.reserve],
            [households[2].id],
        )
        self.assertEqual(summary["reserve_households"], 1)

    def test_select_households_excludes_recently_verified_households(self):
        households = [
            _household(
                "recent",
                "Poorest",
                last_verified_date=date(2026, 7, 2),
            ),
            _household(
                "old",
                "Poorer",
                last_verified_date=date(2026, 6, 1),
            ),
        ]

        result, _ = select_households(
            households,
            exclude_verified_after=date(2026, 6, 30),
        )

        self.assertEqual([row.household.id for row in result.main], ["old"])

    def test_select_households_ignores_households_without_eligible_members(self):
        households = [
            EligibleHousehold(
                id="no-workers",
                code="no-workers",
                wealth_quintile="Poorest",
                eligible_members=[],
            ),
            _household("worker", "Poorer"),
        ]

        result, _ = select_households(households)

        self.assertEqual([row.household.id for row in result.main], ["worker"])

    def test_select_households_backfills_quota_shortage_from_other_categories(self):
        households = [
            _household("female", "Poorest", head_gender="Female"),
            _household("other-1", "Poorer"),
            _household("other-2", "Middle"),
        ]

        result, summary = select_households(households, target_count=3)

        self.assertEqual(len(result.main), 3)
        self.assertEqual(
            [row.household.id for row in result.main],
            ["female", "other-1", "other-2"],
        )
        # "other-2" is only picked up by the backfill loop (there's no youth
        # household to fill that quota slot), so the summary counts must
        # reflect it too, not just the households chosen during quota-fill.
        self.assertEqual(summary["selected_households"], 3)
        self.assertEqual(summary["selected_female_headed_households"], 1)
        self.assertEqual(summary["selected_youth_households"], 0)
        self.assertEqual(summary["selected_other_households"], 2)

    def test_selection_result_expands_selected_households_to_member_rows(self):
        households = [
            _household(
                "household",
                "Poorest",
                eligible_members=[
                    _member("worker-1", age=24),
                    _member("worker-2", age=45),
                ],
            )
        ]

        result, _ = select_households(households)

        self.assertEqual(len(result.member_rows), 2)
        self.assertEqual(
            [row.member.id for row in result.member_rows],
            ["worker-1", "worker-2"],
        )
        self.assertEqual(
            {row.household.id for row in result.member_rows},
            {"household"},
        )

    def test_select_households_treats_ages_18_to_35_as_youth(self):
        households = [
            _household("age-18", "Poorest", eligible_member_age=18),
            _household("age-35", "Poorest", eligible_member_age=35),
            _household("age-36", "Poorest", eligible_member_age=36),
        ]

        result, _ = select_households(households, target_count=3)

        categories = {
            row.household.id: row.category
            for row in result.main
        }
        self.assertEqual(categories["age-18"], CATEGORY_YOUTH)
        self.assertEqual(categories["age-35"], CATEGORY_YOUTH)
        self.assertEqual(categories["age-36"], CATEGORY_OTHER)

    def test_is_truthy_accepts_ubr_boolean_shapes(self):
        for value in (True, 1, "1", "true", "TRUE", "yes", "Y"):
            self.assertTrue(is_truthy(value))
        for value in (False, 0, "0", "false", "no", None, ""):
            self.assertFalse(is_truthy(value))

    def test_select_households_uses_configured_percentages(self):
        households = [
            _household("female", "Poorest", head_gender="Female"),
            _household("youth", "Poorest", eligible_member_age=25),
            _household("other", "Poorest"),
        ]

        with (
            patch.object(HouseholdValidationConfig, "female_headed_percentage", 0),
            patch.object(HouseholdValidationConfig, "youth_percentage", 100),
        ):
            result, _ = select_households(households, target_count=2)

        self.assertEqual(
            [row.category for row in result.main],
            [CATEGORY_YOUTH, CATEGORY_FEMALE_HEADED],
        )

    def test_other_quota_absorbs_the_remainder_when_configured_percentages_total_below_100(self):
        households = (
            [_household(f"female-{i}", "Poorest", head_gender="Female") for i in range(5)]
            + [_household(f"youth-{i}", "Poorest", eligible_member_age=25) for i in range(5)]
            + [_household(f"other-{i}", "Poorest") for i in range(5)]
        )

        with (
            patch.object(HouseholdValidationConfig, "female_headed_percentage", 30),
            patch.object(HouseholdValidationConfig, "youth_percentage", 30),
        ):
            result, summary = select_households(households, target_count=10)

        self.assertEqual(len(result.main), 10)
        self.assertEqual(summary["selected_female_headed_households"], 3)
        self.assertEqual(summary["selected_youth_households"], 3)
        self.assertEqual(summary["selected_other_households"], 4)

    def test_female_and_youth_quotas_are_normalized_when_configured_percentages_exceed_100(self):
        households = (
            [_household(f"female-{i}", "Poorest", head_gender="Female") for i in range(10)]
            + [_household(f"youth-{i}", "Poorest", eligible_member_age=25) for i in range(10)]
        )

        with (
            patch.object(HouseholdValidationConfig, "female_headed_percentage", 70),
            patch.object(HouseholdValidationConfig, "youth_percentage", 60),
        ):
            result, summary = select_households(households, target_count=10)

        self.assertEqual(len(result.main), 10)
        self.assertEqual(summary["selected_other_households"], 0)
        self.assertEqual(
            summary["selected_female_headed_households"] + summary["selected_youth_households"],
            10,
        )


class HouseholdValidationPreviewServiceTest(TestCase):
    def test_generate_and_preview_share_selection_result(self):
        service = _FakeSelectionService(
            [
                _fake_group("group-1", "HH-001", "Female", 30, "Poorest"),
                _fake_group("group-2", "HH-002", "Male", 22, "Poorer"),
                _fake_group("group-3", "HH-003", "Male", 45, "Middle"),
            ]
        )

        filters = {
            "target_count": 2,
        }
        _, summary = service.generate(**filters)
        preview_rows = service.preview(**filters)

        self.assertEqual(summary["total_households"], 3)
        self.assertEqual(summary["total_individuals"], 3)
        self.assertEqual(summary["eligible_households"], 3)
        self.assertEqual(summary["selected_households"], 2)
        self.assertEqual(summary["selected_individuals"], 2)
        self.assertEqual(summary["reserve_households"], 1)
        self.assertEqual(len(preview_rows), 3)
        self.assertEqual(
            len([row for row in preview_rows if row.row_type == ROW_TYPE_MAIN]),
            summary["selected_individuals"],
        )

    def test_generate_totals_cover_full_catchment_before_optional_filters(self):
        catchment_groups = [
            _fake_group("group-1", "HH-001", "Female", 30, "Poorest"),
            _fake_group("group-2", "HH-002", "Male", 22, "Poorer"),
        ]
        service = _FakeSelectionService(catchment_groups)
        filtered_queryset = _FakeQuerySet([catchment_groups[0]])

        with (
            patch.object(
                service,
                "_apply_micro_catchment_scope",
                return_value=_FakeQuerySet(catchment_groups),
            ) as catchment_scope_mock,
            patch.object(
                service,
                "_apply_location_filters",
                return_value=filtered_queryset,
            ),
        ):
            _, summary = service.generate(
                catchment_code="MC01",
                village_codes=["V01"],
                target_count=1,
            )

        catchment_scope_mock.assert_called_once()
        self.assertEqual(summary["total_households"], 2)
        self.assertEqual(summary["total_individuals"], 2)
        self.assertEqual(summary["eligible_households"], 1)
        self.assertEqual(summary["selected_households"], 1)

    def test_preview_rows_include_location_and_validation_fields(self):
        service = _FakeSelectionService(
            [
                _fake_group(
                    "group-1",
                    "HH-001",
                    "Female",
                    30,
                    "Poorest",
                    validation_status="VERIFIED",
                )
            ]
        )

        row = service.preview(target_count=1)[0]

        self.assertEqual(row.group_uuid, "group-1")
        self.assertEqual(row.group_code, "HH-001")
        self.assertEqual(row.head_name, "Head Person")
        self.assertEqual(row.individual_first_name, "Head")
        self.assertEqual(row.region, "Region")
        self.assertEqual(row.district, "District")
        self.assertEqual(row.municipality, "Traditional Authority")
        self.assertEqual(row.village, "Village")
        self.assertEqual(row.wealth_quintile, "Poorest")
        self.assertEqual(row.validation_status, "VERIFIED")


class HotspotAndMicroCatchmentResolutionTest(TestCase):
    @patch("household_validation.services.Hotspot.objects.filter")
    def test_resolve_hotspot_filters_by_uuid_code_and_validity(self, filter_mock):
        filter_mock.return_value.first.return_value = "hotspot-instance"
        service = EligibleHouseholdSelectionService()

        result = service._resolve_hotspot("hotspot-uuid", "HS01")

        self.assertEqual(result, "hotspot-instance")
        self.assertEqual(filter_mock.call_args[0][0], Q(uuid="hotspot-uuid") | Q(code="HS01"))
        self.assertEqual(filter_mock.call_args[1], {"validity_to__isnull": True})

    def test_resolve_hotspot_returns_none_without_identifiers(self):
        service = EligibleHouseholdSelectionService()
        self.assertIsNone(service._resolve_hotspot(None, None))

    @patch("household_validation.services.MicroCatchment.objects.filter")
    def test_resolve_micro_catchment_filters_by_uuid_code_and_validity(self, filter_mock):
        filter_mock.return_value.first.return_value = "catchment-instance"
        service = EligibleHouseholdSelectionService()

        result = service._resolve_micro_catchment("catchment-uuid", "MC01")

        self.assertEqual(result, "catchment-instance")
        self.assertEqual(filter_mock.call_args[0][0], Q(uuid="catchment-uuid") | Q(code="MC01"))
        self.assertEqual(filter_mock.call_args[1], {"validity_to__isnull": True})

    def test_resolve_micro_catchment_returns_none_without_identifiers(self):
        service = EligibleHouseholdSelectionService()
        self.assertIsNone(service._resolve_micro_catchment(None, None))


class LocationFilterHotspotAndCatchmentTest(TestCase):
    def test_scopes_to_hotspot_villages(self):
        service = EligibleHouseholdSelectionService()
        hotspot = SimpleNamespace(
            villages=SimpleNamespace(values_list=lambda *a, **k: ["V01", "V02"]),
        )
        queryset = MagicMock()

        with patch.object(service, "_resolve_hotspot", return_value=hotspot) as resolve_mock:
            result = service._apply_location_filters(queryset, hotspot_code="HS01")

        resolve_mock.assert_called_once_with(None, "HS01")
        queryset.filter.assert_called_once_with(location__code__in=["V01", "V02"])
        self.assertIs(result, queryset.filter.return_value)

    def test_unknown_hotspot_yields_no_households(self):
        service = EligibleHouseholdSelectionService()
        queryset = MagicMock()

        with patch.object(service, "_resolve_hotspot", return_value=None):
            result = service._apply_location_filters(queryset, hotspot_id="missing")

        queryset.none.assert_called_once()
        self.assertIs(result, queryset.none.return_value)

    def test_hotspot_with_no_villages_yields_no_households(self):
        service = EligibleHouseholdSelectionService()
        hotspot = SimpleNamespace(villages=SimpleNamespace(values_list=lambda *a, **k: []))
        queryset = MagicMock()

        with patch.object(service, "_resolve_hotspot", return_value=hotspot):
            result = service._apply_location_filters(queryset, hotspot_code="HS01")

        queryset.none.assert_called_once()
        self.assertIs(result, queryset.none.return_value)

    def test_explicit_village_selection_overrides_hotspot(self):
        service = EligibleHouseholdSelectionService()
        queryset = MagicMock()

        with patch.object(service, "_resolve_hotspot") as resolve_mock:
            service._apply_location_filters(
                queryset,
                village_codes=["V01"],
                hotspot_code="HS01",
            )

        resolve_mock.assert_not_called()

    def test_scopes_to_micro_catchment_hotspot_villages_before_broader_links(self):
        service = EligibleHouseholdSelectionService()
        micro_catchment = SimpleNamespace(
            traditional_authorities=MagicMock(),
            gvhs=MagicMock(),
        )
        queryset = MagicMock()
        hotspot_queryset = MagicMock()
        hotspot_queryset.values_list.return_value.distinct.return_value = ["V01", "V02"]

        with (
            patch.object(service, "_resolve_micro_catchment", return_value=micro_catchment) as resolve_mock,
            patch("household_validation.services.Hotspot.objects.filter", return_value=hotspot_queryset),
        ):
            result = service._apply_location_filters(queryset, catchment_code="MC01")

        resolve_mock.assert_called_once_with(None, "MC01")
        queryset.filter.assert_called_once_with(Q(location__code__in=["V01", "V02"]))
        micro_catchment.gvhs.filter.assert_not_called()
        micro_catchment.traditional_authorities.filter.assert_not_called()
        self.assertIs(result, queryset.filter.return_value)

    def test_micro_catchment_falls_back_to_active_gvhs_without_villages(self):
        service = EligibleHouseholdSelectionService()
        micro_catchment = SimpleNamespace(
            gvhs=MagicMock(),
            traditional_authorities=MagicMock(),
        )
        hotspot_queryset = MagicMock()
        hotspot_queryset.values_list.return_value.distinct.return_value = []
        micro_catchment.gvhs.filter.return_value.values_list.return_value = ["GVH01"]

        with patch(
            "household_validation.services.Hotspot.objects.filter",
            return_value=hotspot_queryset,
        ):
            location_filter = service._micro_catchment_location_filter(micro_catchment)

        self.assertEqual(
            location_filter,
            Q(location__code__in=["GVH01"])
            | Q(location__parent__code__in=["GVH01"]),
        )
        micro_catchment.traditional_authorities.filter.assert_not_called()

    def test_unknown_micro_catchment_yields_no_households(self):
        service = EligibleHouseholdSelectionService()
        queryset = MagicMock()

        with patch.object(service, "_resolve_micro_catchment", return_value=None):
            result = service._apply_location_filters(queryset, catchment_id="missing")

        queryset.none.assert_called_once()
        self.assertIs(result, queryset.none.return_value)

    def test_explicit_ta_selection_is_intersected_with_micro_catchment(self):
        service = EligibleHouseholdSelectionService()
        queryset = MagicMock()
        catchment_queryset = MagicMock()
        queryset.filter.return_value = catchment_queryset
        catchment_filter = Q(location__code__in=["V01"])

        with patch.object(
            service,
            "_micro_catchment_location_filter",
            return_value=catchment_filter,
        ):
            with patch.object(service, "_resolve_micro_catchment", return_value=SimpleNamespace()):
                result = service._apply_location_filters(
                    queryset,
                    ta_codes=["TA01"],
                    catchment_code="MC01",
                )

        queryset.filter.assert_called_once_with(catchment_filter)
        catchment_queryset.filter.assert_called_once_with(
            Q(location__code__in=["TA01"])
            | Q(location__parent__code__in=["TA01"])
            | Q(location__parent__parent__code__in=["TA01"])
        )
        self.assertIs(result, catchment_queryset.filter.return_value)


class CatchmentAndHotspotSelectionIntegrationTest(DjangoTestCase):
    """Prove GraphQL generation and preview exclude out-of-scope households."""

    @classmethod
    def setUpTestData(cls):
        cls.audit_user = User.objects.create(username="household-validation-scope-test")
        cls.district = Location.objects.create(
            code="HV-TEST-DISTRICT",
            name="Validation test district",
            type="R",
            audit_user_id=1,
        )
        cls.ta = Location.objects.create(
            code="HV-TEST-TA",
            name="Validation test TA",
            type="D",
            parent=cls.district,
            audit_user_id=1,
        )
        cls.gvh = Location.objects.create(
            code="HV-TEST-GVH",
            name="Validation test GVH",
            type="W",
            parent=cls.ta,
            audit_user_id=1,
        )
        cls.inside_village = Location.objects.create(
            code="HV-TEST-V-IN",
            name="Inside village",
            type="V",
            parent=cls.gvh,
            audit_user_id=1,
        )
        cls.outside_village = Location.objects.create(
            code="HV-TEST-V-OUT",
            name="Outside village",
            type="V",
            parent=cls.gvh,
            audit_user_id=1,
        )
        cls.catchment = MicroCatchment.objects.create(
            code="HV-TEST-MC",
            name="Validation test micro-catchment",
            district=cls.district,
            audit_user_id=1,
        )
        cls.hotspot = Hotspot.objects.create(
            code="HV-TEST-HS",
            name="Validation test hotspot",
            micro_catchment=cls.catchment,
            audit_user_id=1,
        )
        HotspotVillage.objects.create(
            hotspot=cls.hotspot,
            location=cls.inside_village,
            audit_user_id=1,
        )
        cls.inside_group = cls._create_eligible_household(
            "HV-HH-IN",
            cls.inside_village,
        )
        cls.outside_group = cls._create_eligible_household(
            "HV-HH-OUT",
            cls.outside_village,
        )

    @classmethod
    def _create_eligible_household(cls, code, village):
        group = Group(
            code=code,
            location=village,
            json_ext={
                "form_number": f"FORM-{code}",
                "household_wealth_quintile": "Poorest",
            },
        )
        group.save(user=cls.audit_user)
        individual = Individual(
            first_name=code,
            last_name="Worker",
            dob=date(1990, 1, 1),
            location=village,
            json_ext={"fit_for_work": True, "gender": "Male"},
        )
        individual.save(user=cls.audit_user)
        GroupIndividual.objects.bulk_create(
            [
                GroupIndividual(
                    id=uuid4(),
                    group=group,
                    individual=individual,
                    role=GroupIndividual.Role.HEAD,
                    user_created=cls.audit_user,
                    user_updated=cls.audit_user,
                )
            ]
        )
        return group

    def test_micro_catchment_excludes_district_households_from_generation_and_preview_without_ta(self):
        self._assert_generation_and_preview_scope(
            catchment_code=self.catchment.code,
        )

    def test_hotspot_excludes_households_in_other_villages_in_same_district_from_generation_and_preview(self):
        self._assert_generation_and_preview_scope(
            hotspot_code=self.hotspot.code,
        )

    def _assert_generation_and_preview_scope(self, **scope):
        info = SimpleNamespace(context=SimpleNamespace(user=self.audit_user))
        filters = {
            "district_id": self.district.id,
            "target_count": 10,
            **scope,
        }
        with (
            patch.object(
                Group,
                "get_queryset",
                side_effect=lambda queryset, user: queryset,
            ),
            patch.object(
                GenerateHouseholdValidationListMutation,
                "_validate_user",
            ),
            patch.object(Query, "_check_permissions"),
            patch(
                "household_validation.gql_mutations.HouseholdValidationProjectLookupService.list_projects",
                return_value=[],
            ),
            patch.object(
                EligibleHouseholdSelectionService,
                "_project_names",
                return_value=[],
            ),
        ):
            generated = GenerateHouseholdValidationListMutation.mutate(
                None,
                info,
                **filters,
            )
            preview = Query.resolve_household_validation_preview(
                None,
                info,
                **filters,
            )

        self.assertEqual(generated.total_households, 1)
        self.assertEqual(generated.selected_households, 1)
        self.assertEqual(len(generated.village_breakdown), 1)
        self.assertEqual(
            generated.village_breakdown[0]["village_code"],
            self.inside_village.code,
        )
        self.assertEqual(
            generated.village_breakdown[0]["selected_households"],
            1,
        )
        workbook = load_workbook(
            BytesIO(base64.b64decode(generated.file_base64)),
            read_only=True,
        )
        worksheet = workbook[VALIDATION_LIST_SHEET]
        headers = [cell.value for cell in worksheet[1]]
        group_uuid_column = headers.index("group_uuid") + 1
        generated_group_ids = {
            str(row[group_uuid_column - 1].value)
            for row in worksheet.iter_rows(min_row=2)
        }
        self.assertEqual(generated_group_ids, {str(self.inside_group.id)})
        self.assertNotIn(str(self.outside_group.id), generated_group_ids)

        preview_group_codes = {edge.node.group_code for edge in preview.edges}
        self.assertEqual(preview_group_codes, {self.inside_group.code})
        self.assertNotIn(self.outside_group.code, preview_group_codes)


class _FakeSelectionService(EligibleHouseholdSelectionService):
    def __init__(self, groups):
        super().__init__(user=None)
        self.groups = groups

    def _base_queryset(self):
        return _FakeQuerySet(self.groups)

    def _project_names(self, location):
        return []


class _FakeQuerySet:
    def __init__(self, groups):
        self.groups = groups

    def __iter__(self):
        return iter(self.groups)

    def filter(self, *args, **kwargs):
        return self

    def select_related(self, *args, **kwargs):
        return self

    def prefetch_related(self, *args, **kwargs):
        return self


class _FakeRelatedManager:
    def __init__(self, rows):
        self.rows = rows

    def all(self):
        return self.rows


def _fake_group(
    group_id,
    code,
    head_gender,
    worker_age,
    wealth_quintile,
    validation_status=None,
):
    location = _fake_location_tree()
    individual = SimpleNamespace(
        id=f"{group_id}-individual",
        first_name="Head",
        last_name="Person",
        dob=_dob_for_age(worker_age),
        json_ext={
            "gender": head_gender,
            "fit_for_work": True,
            "household_wealth_quintile": wealth_quintile,
        },
    )
    group_individual = SimpleNamespace(
        id=f"{group_id}-member",
        is_deleted=False,
        role="HEAD",
        recipient_type="PRIMARY",
        individual=individual,
        individual_id=individual.id,
    )
    json_ext = {
        "head_id": individual.id,
        "household_wealth_quintile": wealth_quintile,
    }
    if validation_status:
        json_ext["validation_status"] = validation_status
    return SimpleNamespace(
        id=group_id,
        code=code,
        json_ext=json_ext,
        location=location,
        groupindividuals=_FakeRelatedManager([group_individual]),
    )


def _fake_location_tree():
    region = SimpleNamespace(type="R", name="Region", code="R01", parent=None, id=1)
    district = SimpleNamespace(type="D", name="District", code="D01", parent=region, id=2)
    ta = SimpleNamespace(type="W", name="Traditional Authority", code="TA01", parent=district, id=3)
    return SimpleNamespace(type="V", name="Village", code="V01", parent=ta, id=4)


class ProjectLookupTest(TestCase):
    def test_active_project_statuses_match_mvp_dropdown_scope(self):
        self.assertEqual(ACTIVE_PROJECT_STATUSES, ("PREPARATION", "IN_PROGRESS"))

    def test_project_option_from_project_serializes_project_fields(self):
        project = SimpleNamespace(
            id="project-1",
            name="Road Works",
            status="PREPARATION",
            location_id=12,
        )

        self.assertEqual(
            project_option_from_project(project),
            ProjectOption(
                id="project-1",
                name="Road Works",
                status="PREPARATION",
                location_id="12",
            ),
        )

    def test_project_option_from_project_falls_back_to_location_object(self):
        project = SimpleNamespace(
            id="project-2",
            name="Drainage",
            status="IN_PROGRESS",
            location=SimpleNamespace(id=21),
        )

        self.assertEqual(project_option_from_project(project).location_id, "21")


class GraphQLPermissionTest(TestCase):
    def test_require_permissions_allows_user_with_required_rights(self):
        user = SimpleNamespace(
            id="user-1",
            is_anonymous=False,
            has_perms=lambda perms: perms == ["958001"],
        )

        self.assertIsNone(require_permissions(user, ["958001"]))

    def test_require_permissions_rejects_missing_rights(self):
        user = SimpleNamespace(
            id="user-1",
            is_anonymous=False,
            has_perms=lambda perms: False,
        )

        with self.assertRaises(HouseholdValidationPermissionError):
            require_permissions(user, ["958002"])

    def test_require_permissions_rejects_anonymous_or_missing_user(self):
        anonymous_user = SimpleNamespace(
            id=None,
            is_anonymous=True,
            has_perms=lambda perms: True,
        )

        with self.assertRaises(HouseholdValidationPermissionError):
            require_permissions(anonymous_user, ["958001"])
        with self.assertRaises(HouseholdValidationPermissionError):
            require_permissions(None, ["958001"])


class ValidationUploadParserTest(TestCase):
    def test_parse_validation_workbook_resolves_project_name_to_hidden_project_id(self):
        workbook = self._upload_workbook()
        worksheet = workbook[VALIDATION_LIST_SHEET]
        project_column = EXCEL_COLUMNS.index("project") + 1
        verified_column = EXCEL_COLUMNS.index("verified") + 1
        primary_worker_column = EXCEL_COLUMNS.index("primary_worker") + 1
        worksheet.cell(row=2, column=project_column, value="Road Works")
        worksheet.cell(row=2, column=verified_column, value="YES")
        worksheet.cell(row=2, column=primary_worker_column, value="YES")

        parsed = parse_validation_workbook(self._workbook_bytes(workbook))

        self.assertEqual(parsed.errors, [])
        self.assertEqual(parsed.rows_read, 1)
        self.assertEqual(parsed.rows[0].project_id, "project-1")
        self.assertEqual(parsed.rows[0].verified, True)
        self.assertEqual(parsed.rows[0].primary_worker, True)
        self.assertEqual(PROJECT_SELECTION_TYPE_INTENT, "INTENT")

    def test_parse_validation_workbook_reports_missing_required_columns(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = VALIDATION_LIST_SHEET
        worksheet.append(["batch_id", "group_uuid"])

        parsed = parse_validation_workbook(self._workbook_bytes(workbook))

        self.assertEqual(parsed.rows, [])
        self.assertTrue(parsed.errors[0].startswith("Missing required columns:"))
        self.assertIn("member_uuid", parsed.errors[0])

    def test_parse_validation_workbook_accepts_previous_schema(self):
        workbook = self._upload_workbook()
        worksheet = workbook[VALIDATION_LIST_SHEET]
        for column_number in sorted(
            (
                EXCEL_COLUMNS.index(column) + 1
                for column in OPTIONAL_UPLOAD_COLUMNS
            ),
            reverse=True,
        ):
            worksheet.delete_cols(column_number)
        worksheet.cell(
            row=1,
            column=worksheet.max_column + 1,
            value="current_recipient_type",
        )

        parsed = parse_validation_workbook(self._workbook_bytes(workbook))

        self.assertEqual(parsed.errors, [])
        self.assertEqual(parsed.rows_read, 1)
        for column in OPTIONAL_UPLOAD_COLUMNS:
            self.assertIsNone(parsed.rows[0].values[column])

    def test_parse_validation_workbook_preserves_primary_worker_no_and_blank(self):
        workbook = self._upload_workbook()
        worksheet = workbook[VALIDATION_LIST_SHEET]
        primary_worker_column = EXCEL_COLUMNS.index("primary_worker") + 1
        worksheet.cell(row=2, column=primary_worker_column, value="NO")
        worksheet.append(
            [
                "batch-1" if column == "batch_id"
                else "group-2" if column == "group_uuid"
                else "member-2" if column == "member_uuid"
                else None
                for column in EXCEL_COLUMNS
            ]
        )

        parsed = parse_validation_workbook(self._workbook_bytes(workbook))

        self.assertEqual(parsed.errors, [])
        self.assertEqual(parsed.rows[0].primary_worker, False)
        self.assertIsNone(parsed.rows[1].primary_worker)

    def test_parse_validation_workbook_rejects_invalid_editable_values(self):
        workbook = self._upload_workbook()
        worksheet = workbook[VALIDATION_LIST_SHEET]
        verified_column = EXCEL_COLUMNS.index("verified") + 1
        primary_worker_column = EXCEL_COLUMNS.index("primary_worker") + 1
        worksheet.cell(row=2, column=verified_column, value="MAYBE")
        worksheet.cell(row=2, column=primary_worker_column, value="LATER")

        parsed = parse_validation_workbook(self._workbook_bytes(workbook))

        self.assertEqual(parsed.rows, [])
        self.assertIn("Row 2: verified must be YES or NO", parsed.errors)
        self.assertIn("Row 2: primary_worker must be YES or NO", parsed.errors)
        self.assertEqual(parsed.error_row_numbers, frozenset({2}))

    def test_parse_validation_workbook_rejects_unknown_project_names(self):
        workbook = self._upload_workbook()
        worksheet = workbook[VALIDATION_LIST_SHEET]
        project_column = EXCEL_COLUMNS.index("project") + 1
        worksheet.cell(row=2, column=project_column, value="Unknown Project")

        parsed = parse_validation_workbook(self._workbook_bytes(workbook))

        self.assertEqual(parsed.rows, [])
        self.assertIn("Row 2: project is not in the project options", parsed.errors)

    def test_parse_validation_workbook_rejects_hidden_project_id_without_project(self):
        workbook = self._upload_workbook()
        worksheet = workbook[VALIDATION_LIST_SHEET]
        project_id_column = EXCEL_COLUMNS.index("project_id") + 1
        worksheet.cell(row=2, column=project_id_column, value="project-1")

        parsed = parse_validation_workbook(self._workbook_bytes(workbook))

        self.assertEqual(parsed.rows, [])
        self.assertIn("Row 2: project_id cannot be set without project", parsed.errors)

    def test_parse_validation_workbook_rejects_hidden_project_id_mismatch(self):
        workbook = self._upload_workbook()
        worksheet = workbook[VALIDATION_LIST_SHEET]
        project_column = EXCEL_COLUMNS.index("project") + 1
        project_id_column = EXCEL_COLUMNS.index("project_id") + 1
        worksheet.cell(row=2, column=project_column, value="Road Works")
        worksheet.cell(row=2, column=project_id_column, value="different-project")

        parsed = parse_validation_workbook(self._workbook_bytes(workbook))

        self.assertEqual(parsed.rows, [])
        self.assertIn("Row 2: project is not in the project options", parsed.errors)

    def test_parse_validation_workbook_resolves_duplicate_project_name_labels(self):
        workbook = self._upload_workbook()
        worksheet = workbook[VALIDATION_LIST_SHEET]
        project_options = workbook[PROJECT_OPTIONS_SHEET]
        project_options.cell(row=1, column=3, value="project_label")
        project_options.cell(row=2, column=3, value="Road Works (project-1)")
        project_options.append(["project-2", "Road Works", "Road Works (project-2)"])
        project_column = EXCEL_COLUMNS.index("project") + 1
        worksheet.cell(row=2, column=project_column, value="Road Works (project-2)")

        parsed = parse_validation_workbook(self._workbook_bytes(workbook))

        self.assertEqual(parsed.errors, [])
        self.assertEqual(parsed.rows[0].project_id, "project-2")
        self.assertEqual(parsed.rows[0].project_name, "Road Works")

    def _upload_workbook(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = VALIDATION_LIST_SHEET
        worksheet.append(EXCEL_COLUMNS)
        values = {column: None for column in EXCEL_COLUMNS}
        values.update(
            {
                "batch_id": "batch-1",
                "group_uuid": "group-1",
                "member_uuid": "member-1",
                "District": "District",
                "TA": "Traditional Authority",
                "GVH": "Group Village Head",
                "Village": "Village",
            }
        )
        worksheet.append([values[column] for column in EXCEL_COLUMNS])
        project_options = workbook.create_sheet(PROJECT_OPTIONS_SHEET)
        project_options.append(["project_id", "project"])
        project_options.append(["project-1", "Road Works"])
        return workbook

    def _workbook_bytes(self, workbook):
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()


class ExcelValidationListExporterTest(TestCase):
    def test_export_workbook_writes_headers_and_member_rows(self):
        result = self._selection_result()

        workbook = ExcelValidationListExporter(result, batch_id="batch-1").export_workbook()
        worksheet = workbook["Validation List"]

        self.assertEqual(
            [worksheet.cell(row=1, column=index).value for index in range(1, len(EXCEL_COLUMNS) + 1)],
            EXCEL_COLUMNS,
        )
        self.assertIn("primary_worker", EXCEL_COLUMNS)
        self.assertNotIn("head", EXCEL_COLUMNS)
        self.assertNotIn("validation_date", EXCEL_COLUMNS)
        self.assertNotIn("current_recipient_type", EXCEL_COLUMNS)
        form_number_index = EXCEL_COLUMNS.index("form_number")
        self.assertEqual(
            EXCEL_COLUMNS[form_number_index:form_number_index + 7],
            [
                "form_number",
                "member_name",
                "relationship",
                "member_dob",
                "national_id",
                "primary_worker",
                "verified",
            ],
        )
        self.assertNotIn("participant", EXCEL_COLUMNS)
        self.assertNotIn("group_code", EXCEL_COLUMNS)
        self.assertEqual(worksheet["A2"].value, "batch-1")
        self.assertEqual(worksheet["B2"].value, ROW_TYPE_MAIN)
        self.assertEqual(self._value(worksheet, "District"), "District")
        self.assertEqual(self._value(worksheet, "TA"), "Traditional Authority")
        self.assertEqual(self._value(worksheet, "GVH"), "Group Village Head")
        self.assertEqual(self._value(worksheet, "Village"), "Village")
        self.assertEqual(self._value(worksheet, "form_number"), "FORM-001")
        self.assertEqual(self._value(worksheet, "group_uuid"), "group-1")
        self.assertEqual(self._value(worksheet, "member_uuid"), "member-1")
        self.assertEqual(self._value(worksheet, "member_name"), "Ada Worker")
        self.assertEqual(self._value(worksheet, "national_id"), "NAT-001")
        self.assertEqual(self._value(worksheet, "marital_status"), "Married")
        self.assertEqual(self._value(worksheet, "disability"), "Not disabled")
        self.assertEqual(self._value(worksheet, "fit_for_work"), "YES")
        self.assertEqual(self._value(worksheet, "relationship"), "HEAD")
        self.assertEqual(self._value(worksheet, "pmt_score"), -1.234)
        self.assertEqual(
            self._value(worksheet, "household_wealth_quintile"),
            "Poorest",
        )

        group = result.main[0].household.source
        self.assertEqual(group.code, "HH-001")

    def test_export_workbook_writes_exact_relationship_roles(self):
        workbook = ExcelValidationListExporter(
            self._selection_result(member_count=3),
            batch_id="batch-1",
        ).export_workbook()
        worksheet = workbook["Validation List"]
        relationship_column = EXCEL_COLUMNS.index("relationship") + 1

        self.assertEqual(worksheet.cell(2, relationship_column).value, "HEAD")
        self.assertEqual(worksheet.cell(3, relationship_column).value, "SPOUSE")
        self.assertEqual(worksheet.cell(4, relationship_column).value, "SON")

    def test_export_workbook_has_hidden_project_id_and_dropdowns(self):
        result = self._selection_result()
        projects = [SimpleNamespace(id="project-1", name="Road Works, Phase 1")]

        workbook = ExcelValidationListExporter(
            result,
            batch_id="batch-1",
            projects=projects,
        ).export_workbook()
        worksheet = workbook["Validation List"]
        project_options = workbook[PROJECT_OPTIONS_SHEET]
        project_id_letter = worksheet.cell(1, EXCEL_COLUMNS.index("project_id") + 1).column_letter

        self.assertTrue(worksheet.column_dimensions[project_id_letter].hidden)
        self.assertEqual(project_options.sheet_state, "hidden")
        self.assertEqual(project_options["A2"].value, "project-1")
        self.assertEqual(project_options["B2"].value, "Road Works, Phase 1")
        self.assertEqual(project_options["C2"].value, "Road Works, Phase 1")
        formulas = {validation.formula1 for validation in worksheet.data_validations.dataValidation}
        self.assertIn('"YES,NO"', formulas)
        self.assertIn("'Project Options'!$C$2:$C$2", formulas)

    def test_export_workbook_disambiguates_duplicate_project_names(self):
        result = self._selection_result()
        projects = [
            SimpleNamespace(id="project-1", name="Road Works"),
            SimpleNamespace(id="project-2", name="Road Works"),
        ]

        workbook = ExcelValidationListExporter(
            result,
            batch_id="batch-1",
            projects=projects,
        ).export_workbook()
        project_options = workbook[PROJECT_OPTIONS_SHEET]

        self.assertEqual(project_options["C2"].value, "Road Works (project-1)")
        self.assertEqual(project_options["C3"].value, "Road Works (project-2)")

    def test_export_workbook_locks_structural_cells_and_unlocks_field_inputs(self):
        workbook = ExcelValidationListExporter(
            self._selection_result(),
            batch_id="batch-1",
        ).export_workbook()
        worksheet = workbook["Validation List"]

        self.assertTrue(worksheet.protection.sheet)
        for column in (
            "group_uuid",
            "relationship",
            "household_wealth_quintile",
        ):
            self.assertTrue(self._cell(worksheet, column).protection.locked)
        for column in (
            "national_id",
            "primary_worker",
            "verified",
            "validation_notes",
        ):
            self.assertFalse(self._cell(worksheet, column).protection.locked)

    def test_export_workbook_writes_one_row_per_selected_eligible_member(self):
        result = self._selection_result(member_count=2)

        workbook = ExcelValidationListExporter(result, batch_id="batch-1").export_workbook()
        worksheet = workbook["Validation List"]

        self.assertEqual(worksheet.max_row, 3)
        self.assertEqual(self._value(worksheet, "member_uuid", 2), "member-1")
        self.assertEqual(self._value(worksheet, "member_uuid", 3), "member-2")

    def test_export_workbook_preserves_identifier_columns_as_text(self):
        result = self._selection_result()
        member = result.main[0].household.eligible_members[0]
        individual = member.source.individual
        individual.json_ext["form_number"] = "001234"
        individual.json_ext["national_id"] = "00123456"

        exported = ExcelValidationListExporter(
            result,
            batch_id="batch-1",
        ).export_bytes()
        worksheet = load_workbook(BytesIO(exported))["Validation List"]

        for column, expected in (
            ("form_number", "001234"),
            ("national_id", "00123456"),
        ):
            cell = self._cell(worksheet, column)
            self.assertEqual(cell.value, expected)
            self.assertEqual(cell.data_type, "s")
            self.assertEqual(cell.number_format, "@")

    def test_export_workbook_writes_stored_primary_worker_values(self):
        result = self._selection_result(member_count=3)
        stored_values = (True, False, None)
        for selected_member, stored_value in zip(
            result.main[0].household.eligible_members,
            stored_values,
        ):
            selected_member.source.json_ext = (
                {"primary_worker": stored_value}
                if stored_value is not None
                else {}
            )

        worksheet = ExcelValidationListExporter(
            result,
            batch_id="batch-1",
        ).export_workbook()["Validation List"]

        self.assertEqual(self._value(worksheet, "primary_worker", 2), "YES")
        self.assertEqual(self._value(worksheet, "primary_worker", 3), "NO")
        self.assertIsNone(self._value(worksheet, "primary_worker", 4))

    def test_primary_worker_suggests_yes_for_primary_recipient_when_unverified(self):
        exporter = ExcelValidationListExporter(
            SelectionResult(main=[], reserve=[]), batch_id="batch-1",
        )
        group_individual = SimpleNamespace(json_ext={}, recipient_type="PRIMARY")

        self.assertEqual(exporter._primary_worker(group_individual), "YES")

    def test_primary_worker_leaves_non_primary_recipient_blank_when_unverified(self):
        exporter = ExcelValidationListExporter(
            SelectionResult(main=[], reserve=[]), batch_id="batch-1",
        )
        secondary = SimpleNamespace(json_ext={}, recipient_type="SECONDARY")
        unset = SimpleNamespace(json_ext={}, recipient_type=None)

        self.assertIsNone(exporter._primary_worker(secondary))
        self.assertIsNone(exporter._primary_worker(unset))

    def test_primary_worker_prefers_verified_value_over_recipient_type_suggestion(self):
        exporter = ExcelValidationListExporter(
            SelectionResult(main=[], reserve=[]), batch_id="batch-1",
        )
        confirmed_no = SimpleNamespace(json_ext={"primary_worker": False}, recipient_type="PRIMARY")
        confirmed_yes = SimpleNamespace(json_ext={"primary_worker": True}, recipient_type="SECONDARY")

        self.assertEqual(exporter._primary_worker(confirmed_no), "NO")
        self.assertEqual(exporter._primary_worker(confirmed_yes), "YES")

    def test_export_workbook_resolves_micro_catchment_from_gvh_link(self):
        micro_catchment = SimpleNamespace(name="Catchment A", code="MC-A")
        link = SimpleNamespace(micro_catchment=micro_catchment)
        gvh_manager = MagicMock()
        gvh_manager.filter.return_value.select_related.return_value.first.return_value = link

        district = SimpleNamespace(type="R", name="District", code="D01", parent=None)
        ta = SimpleNamespace(
            type="D", name="Traditional Authority", code="TA01", parent=district, id=2,
        )
        gvh = SimpleNamespace(
            type="W", name="Group Village Head", code="GVH01", parent=ta, id=3,
            micro_catchments_gvh=gvh_manager,
        )
        village = SimpleNamespace(type="V", name="Village", code="V01", parent=gvh)

        result = self._selection_result(location=village)
        worksheet = ExcelValidationListExporter(
            result,
            batch_id="batch-1",
        ).export_workbook()["Validation List"]

        self.assertEqual(self._value(worksheet, MICRO_CATCHMENT_COLUMN), "Catchment A")
        gvh_manager.filter.assert_called_once_with(
            validity_to__isnull=True,
            micro_catchment__validity_to__isnull=True,
        )

    def test_export_workbook_falls_back_to_ta_micro_catchment_link(self):
        micro_catchment = SimpleNamespace(name=None, code="MC-B")
        link = SimpleNamespace(micro_catchment=micro_catchment)
        ta_manager = MagicMock()
        ta_manager.filter.return_value.select_related.return_value.first.return_value = link
        gvh_manager = MagicMock()
        gvh_manager.filter.return_value.select_related.return_value.first.return_value = None

        district = SimpleNamespace(type="R", name="District", code="D01", parent=None)
        ta = SimpleNamespace(
            type="D", name="Traditional Authority", code="TA01", parent=district, id=2,
            micro_catchments_ta=ta_manager,
        )
        gvh = SimpleNamespace(
            type="W", name="Group Village Head", code="GVH01", parent=ta, id=3,
            micro_catchments_gvh=gvh_manager,
        )
        village = SimpleNamespace(type="V", name="Village", code="V01", parent=gvh)

        result = self._selection_result(location=village)
        worksheet = ExcelValidationListExporter(
            result,
            batch_id="batch-1",
        ).export_workbook()["Validation List"]

        # No name on the micro-catchment, so the code is used as the label.
        self.assertEqual(self._value(worksheet, MICRO_CATCHMENT_COLUMN), "MC-B")

    def test_export_workbook_leaves_micro_catchment_blank_without_a_link(self):
        result = self._selection_result()

        worksheet = ExcelValidationListExporter(
            result,
            batch_id="batch-1",
        ).export_workbook()["Validation List"]

        self.assertIsNone(self._value(worksheet, MICRO_CATCHMENT_COLUMN))

    def test_export_workbook_resolves_hotspot_from_village_link(self):
        hotspot = SimpleNamespace(name="Hotspot A", code="HS-A")
        link = SimpleNamespace(hotspot=hotspot)
        village_manager = MagicMock()
        village_manager.filter.return_value.select_related.return_value.first.return_value = link

        district = SimpleNamespace(type="R", name="District", code="D01", parent=None)
        ta = SimpleNamespace(type="D", name="Traditional Authority", code="TA01", parent=district)
        gvh = SimpleNamespace(type="W", name="Group Village Head", code="GVH01", parent=ta)
        village = SimpleNamespace(
            type="V", name="Village", code="V01", parent=gvh, id=4,
            hotspot_links=village_manager,
        )

        result = self._selection_result(location=village)
        worksheet = ExcelValidationListExporter(
            result,
            batch_id="batch-1",
        ).export_workbook()["Validation List"]

        self.assertEqual(self._value(worksheet, HOTSPOT_COLUMN), "Hotspot A")
        village_manager.filter.assert_called_once_with(
            validity_to__isnull=True,
            hotspot__validity_to__isnull=True,
        )

    def test_export_workbook_leaves_hotspot_blank_without_a_link(self):
        result = self._selection_result()

        worksheet = ExcelValidationListExporter(
            result,
            batch_id="batch-1",
        ).export_workbook()["Validation List"]

        self.assertIsNone(self._value(worksheet, HOTSPOT_COLUMN))

    def _selection_result(self, member_count=1, location=None):
        location = location or self._location_tree()
        group = SimpleNamespace(id="group-1", code="HH-001", location=location)
        selected_members = []
        for index in range(1, member_count + 1):
            individual = SimpleNamespace(
                first_name="Ada" if index == 1 else "Grace",
                last_name="Worker",
                json_ext={
                    "form_number": "FORM-001",
                    "national_id": f"NAT-{index:03d}",
                    "marital_status": "Married" if index == 1 else "Never married",
                    "disability": "Not disabled",
                },
            )
            group_individual = SimpleNamespace(individual=individual)
            selected_members.append(
                EligibleMember(
                    id=f"member-{index}",
                    gender="Female",
                    dob=date(1990, 1, index),
                    fit_for_work=True,
                    role=("HEAD" if index == 1 else "SPOUSE" if index == 2 else "SON"),
                    recipient_type="PRIMARY" if index == 1 else None,
                    source=group_individual,
                )
            )
        household = EligibleHousehold(
            id="group-1",
            code="HH-001",
            wealth_quintile="Poorest",
            pmt_score=-1.234,
            head=selected_members[0],
            eligible_members=selected_members,
            source=group,
        )
        return SelectionResult(
            main=[
                SelectedHousehold(
                    household=household,
                    category=CATEGORY_FEMALE_HEADED,
                    row_type=ROW_TYPE_MAIN,
                )
            ],
            reserve=[],
        )

    def _cell(self, worksheet, column_name, row_number=2):
        return worksheet.cell(row_number, EXCEL_COLUMNS.index(column_name) + 1)

    def _value(self, worksheet, column_name, row_number=2):
        return self._cell(worksheet, column_name, row_number).value

    def _location_tree(self):
        district = SimpleNamespace(type="R", name="District", code="D01", parent=None)
        ta = SimpleNamespace(type="D", name="Traditional Authority", code="TA01", parent=district)
        gvh = SimpleNamespace(type="W", name="Group Village Head", code="GVH01", parent=ta)
        return SimpleNamespace(type="V", name="Village", code="V01", parent=gvh)


class UploadHardeningTest(TestCase):
    @staticmethod
    def _uploaded_row(row_number, member_uuid, primary_worker):
        return UploadedValidationRow(
            row_number=row_number,
            values={
                "group_uuid": "group-1",
                "member_uuid": member_uuid,
                "form_number": "FORM-001",
            },
            verified=True,
            primary_worker=primary_worker,
            validation_date=None,
            project_name=None,
            project_id=None,
            notes=None,
        )

    @staticmethod
    def _group_with_primary_workers(*primary_worker_values):
        members = [
            SimpleNamespace(
                individual_id=f"member-{index}",
                is_deleted=False,
                json_ext={"primary_worker": primary_worker},
            )
            for index, primary_worker in enumerate(primary_worker_values, start=1)
        ]
        return SimpleNamespace(
            id="group-1",
            code="HH-001",
            _prefetched_objects_cache={"groupindividuals": members},
        )

    def test_primary_worker_preflight_rejects_existing_and_new_primary_worker(self):
        group = self._group_with_primary_workers(True, False)
        rows = [
            self._uploaded_row(2, "member-1", None),
            self._uploaded_row(3, "member-2", True),
        ]
        service = HouseholdValidationUploadService()

        with (
            patch.object(service, "_group", return_value=group),
            patch.object(service, "_structural_errors", return_value=[]),
        ):
            rejections = service._primary_worker_rejections(rows)

        self.assertEqual(set(rejections), {"group-1"})
        self.assertEqual(rejections["group-1"]["row_count"], 2)
        self.assertEqual(rejections["group-1"]["row_numbers"], {2, 3})

    def test_primary_worker_preflight_accepts_explicit_worker_transfer(self):
        group = self._group_with_primary_workers(True, False)
        rows = [
            self._uploaded_row(2, "member-1", False),
            self._uploaded_row(3, "member-2", True),
        ]
        service = HouseholdValidationUploadService()

        with (
            patch.object(service, "_group", return_value=group),
            patch.object(service, "_structural_errors", return_value=[]),
        ):
            rejections = service._primary_worker_rejections(rows)

        self.assertEqual(rejections, {})

    def test_primary_worker_preflight_ignores_unverified_households(self):
        group = self._group_with_primary_workers(True, False)
        rows = [
            UploadedValidationRow(
                row_number=2,
                values={
                    "group_uuid": "group-1",
                    "member_uuid": "member-2",
                    "form_number": "FORM-001",
                },
                verified=None,
                primary_worker=True,
                validation_date=None,
                project_name=None,
                project_id=None,
                notes=None,
            ),
        ]
        service = HouseholdValidationUploadService()

        with patch.object(service, "_group", return_value=group):
            rejections = service._primary_worker_rejections(
                rows,
                eligible_group_keys=set(),
            )

        self.assertEqual(rejections, {})

    @patch("household_validation.services.parse_validation_workbook")
    def test_dry_run_reports_unique_rejected_household_without_writes(
        self,
        parse_workbook_mock,
    ):
        rows = [
            self._uploaded_row(2, "member-1", True),
            self._uploaded_row(3, "member-2", True),
        ]
        parse_workbook_mock.return_value = SimpleNamespace(
            rows=rows,
            rows_read=2,
            errors=[],
        )
        group = self._group_with_primary_workers(False, False)
        service = HouseholdValidationUploadService()

        with (
            patch.object(service, "_group", return_value=group),
            patch.object(service, "_structural_errors", return_value=[]),
            patch.object(service, "_get_or_create_batch") as create_batch_mock,
        ):
            result = service.upload(b"workbook", dry_run=True)

        self.assertEqual(result["households_with_multiple_primary_workers"], 1)
        self.assertEqual(result["errors"], 2)
        self.assertEqual(result["error_messages"], [])
        create_batch_mock.assert_not_called()

    @patch("household_validation.services.parse_validation_workbook")
    def test_dry_run_does_not_double_count_parse_and_household_rejection(
        self,
        parse_workbook_mock,
    ):
        rows = [
            self._uploaded_row(2, "member-1", True),
            self._uploaded_row(3, "member-2", True),
        ]
        parse_workbook_mock.return_value = SimpleNamespace(
            rows=rows,
            rows_read=2,
            errors=["Row 2: invalid value"],
            error_row_numbers=frozenset({2}),
        )
        group = self._group_with_primary_workers(False, False)
        service = HouseholdValidationUploadService()

        with (
            patch.object(service, "_group", return_value=group),
            patch.object(service, "_structural_errors", return_value=[]),
        ):
            result = service.upload(b"workbook", dry_run=True)

        self.assertEqual(result["errors"], 2)

    def test_primary_worker_rejection_explains_reason(self):
        service = HouseholdValidationUploadService()
        uploaded_row = self._uploaded_row(2, "member-1", True)

        with (
            patch.object(service, "_group", return_value=None),
            patch.object(service, "_group_individual", return_value=None),
            patch.object(service, "_project", return_value=None),
            patch.object(service, "_save_batch_row") as save_row_mock,
        ):
            service._save_primary_worker_rejection(
                uploaded_row,
                batch=MagicMock(),
            )

        self.assertEqual(
            save_row_mock.call_args.kwargs["error_message"],
            "household has more than one primary worker",
        )
        self.assertEqual(
            save_row_mock.call_args.kwargs["error_code"],
            "MULTIPLE_PRIMARY_WORKERS",
        )

    @patch("household_validation.services.HouseholdValidationBatchRow")
    def test_saved_rows_are_scoped_to_current_upload_attempt(self, batch_row_mock):
        service = HouseholdValidationUploadService()
        service._upload_attempt_id = uuid4()

        service._save_batch_row(
            batch=MagicMock(),
            uploaded_row=self._uploaded_row(2, "member-1", True),
        )

        self.assertEqual(
            batch_row_mock.call_args.kwargs["upload_attempt_id"],
            service._upload_attempt_id,
        )

    @patch("household_validation.services.parse_validation_workbook")
    def test_upload_rejects_every_row_without_partial_household_updates(
        self,
        parse_workbook_mock,
    ):
        rows = [
            self._uploaded_row(2, "member-1", True),
            self._uploaded_row(3, "member-2", True),
        ]
        parse_workbook_mock.return_value = SimpleNamespace(
            rows=rows,
            rows_read=2,
            errors=[],
        )
        group = self._group_with_primary_workers(False, False)
        batch = MagicMock()
        batch.id = "batch-1"
        service = HouseholdValidationUploadService()

        with (
            patch.object(service, "_group", return_value=group),
            patch.object(service, "_structural_errors", return_value=[]),
            patch.object(service, "_get_or_create_batch", return_value=batch),
            patch.object(
                service,
                "_save_primary_worker_rejection",
            ) as reject_mock,
            patch.object(service, "_apply_row") as apply_row_mock,
            patch(
                "household_validation.services.transaction.atomic",
                return_value=nullcontext(),
            ),
        ):
            result = service.upload(b"workbook")

        self.assertEqual(result["households_with_multiple_primary_workers"], 1)
        self.assertEqual(result["households_verified"], 0)
        self.assertEqual(result["households_not_verified"], 0)
        self.assertEqual(result["batch_id"], "batch-1")
        self.assertEqual(result["upload_attempt_id"], service._upload_attempt_id)
        self.assertIsNotNone(result["upload_attempt_id"])
        self.assertEqual(reject_mock.call_count, 2)
        apply_row_mock.assert_not_called()

    @patch("household_validation.services.GroupIndividualService")
    def test_primary_worker_update_preserves_json_and_recipient_type(
        self,
        service_class_mock,
    ):
        group_individual = SimpleNamespace(
            id="membership-1",
            group_id="group-1",
            recipient_type="SECONDARY",
            json_ext={"existing": "value"},
        )

        service = HouseholdValidationUploadService()
        service._apply_primary_worker(group_individual, True)
        service._apply_primary_worker(group_individual, False)

        service_class_mock.return_value.update.assert_has_calls(
            [
                call(
                    {
                        "id": "membership-1",
                        "group_id": "group-1",
                        "json_ext": {
                            "existing": "value",
                            "primary_worker": True,
                        },
                    }
                ),
                call(
                    {
                        "id": "membership-1",
                        "group_id": "group-1",
                        "json_ext": {
                            "existing": "value",
                            "primary_worker": False,
                        },
                    }
                ),
            ]
        )
        self.assertEqual(group_individual.recipient_type, "SECONDARY")

    @patch("household_validation.services.IndividualService")
    def test_national_id_update_preserves_individual_json(self, service_class_mock):
        individual = SimpleNamespace(
            id="individual-1",
            json_ext={"existing": "value", "national_id": "OLD-001"},
        )
        group_individual = SimpleNamespace(individual=individual)
        service = HouseholdValidationUploadService()

        service._apply_national_id(group_individual, " NEW-002 ")

        service_class_mock.return_value.update.assert_called_once_with(
            {
                "id": "individual-1",
                "json_ext": {
                    "existing": "value",
                    "national_id": "NEW-002",
                },
            }
        )

    def test_apply_row_counts_changed_national_id_as_participant_update(self):
        group = SimpleNamespace(id="group-1")
        group_individual = SimpleNamespace(
            individual=SimpleNamespace(
                id="individual-1",
                json_ext={"national_id": "OLD-001"},
            ),
        )
        uploaded_row = UploadedValidationRow(
            row_number=2,
            values={
                "group_uuid": "group-1",
                "member_uuid": "member-1",
                "form_number": "FORM-001",
                "national_id": "NEW-002",
            },
            verified=None,
            primary_worker=None,
            validation_date=None,
            project_name=None,
            project_id=None,
            notes=None,
        )
        service = HouseholdValidationUploadService()

        with (
            patch.object(
                service,
                "_resolve_row",
                return_value=([], group, group_individual, None),
            ),
            patch.object(service, "_apply_national_id") as apply_national_id_mock,
            patch.object(service, "_save_batch_row") as save_row_mock,
        ):
            errors, participant_updated = service._apply_row(
                uploaded_row,
                batch=MagicMock(),
                upload_date=date.today(),
                uploaded_at=datetime.now(),
                allow_participant_update=True,
            )

        self.assertEqual(errors, [])
        self.assertTrue(participant_updated)
        apply_national_id_mock.assert_called_once_with(
            group_individual,
            "NEW-002",
        )
        self.assertEqual(
            save_row_mock.call_args.kwargs["status"],
            HouseholdValidationBatchRow.Status.APPLIED,
        )

    def test_apply_row_updates_only_changed_primary_worker_value(self):
        group = SimpleNamespace(id="group-1")
        group_individual = SimpleNamespace(
            individual_id="member-1",
            json_ext={"primary_worker": True},
        )
        service = HouseholdValidationUploadService()

        def uploaded_row(primary_worker):
            return UploadedValidationRow(
                row_number=2,
                values={
                    "group_uuid": "group-1",
                    "member_uuid": "member-1",
                    "form_number": "FORM-001",
                },
                verified=None,
                primary_worker=primary_worker,
                validation_date=None,
                project_name=None,
                project_id=None,
                notes=None,
            )

        with (
            patch.object(service, "_group", return_value=group),
            patch.object(
                service,
                "_group_individual",
                return_value=group_individual,
            ),
            patch.object(service, "_project", return_value=None),
            patch.object(service, "_structural_errors", return_value=[]),
            patch.object(service, "_apply_primary_worker") as apply_worker_mock,
            patch.object(service, "_save_batch_row") as save_row_mock,
        ):
            unchanged_errors, unchanged = service._apply_row(
                uploaded_row(True),
                batch=MagicMock(),
                upload_date=date.today(),
                uploaded_at=datetime.now(),
            )
            blank_errors, blank = service._apply_row(
                uploaded_row(None),
                batch=MagicMock(),
                upload_date=date.today(),
                uploaded_at=datetime.now(),
            )
            changed_errors, changed = service._apply_row(
                uploaded_row(False),
                batch=MagicMock(),
                upload_date=date.today(),
                uploaded_at=datetime.now(),
                allow_participant_update=True,
            )

        self.assertEqual(unchanged_errors, [])
        self.assertFalse(unchanged)
        self.assertEqual(blank_errors, [])
        self.assertFalse(blank)
        self.assertEqual(changed_errors, [])
        self.assertTrue(changed)
        apply_worker_mock.assert_called_once_with(group_individual, False)
        self.assertEqual(
            [
                recorded_call.kwargs["status"]
                for recorded_call in save_row_mock.call_args_list
            ],
            [
                HouseholdValidationBatchRow.Status.SKIPPED,
                HouseholdValidationBatchRow.Status.SKIPPED,
                HouseholdValidationBatchRow.Status.APPLIED,
            ],
        )

    def test_apply_row_does_not_update_participant_without_accepted_validation(self):
        group = SimpleNamespace(id="group-1")
        group_individual = SimpleNamespace(
            individual_id="member-1",
            json_ext={"primary_worker": False},
        )
        uploaded_row = UploadedValidationRow(
            row_number=2,
            values={
                "group_uuid": "group-1",
                "member_uuid": "member-1",
                "form_number": "FORM-001",
            },
            verified=None,
            primary_worker=True,
            validation_date=None,
            project_name=None,
            project_id=None,
            notes=None,
        )
        service = HouseholdValidationUploadService()

        with (
            patch.object(service, "_group", return_value=group),
            patch.object(service, "_group_individual", return_value=group_individual),
            patch.object(service, "_project", return_value=None),
            patch.object(service, "_structural_errors", return_value=[]),
            patch.object(service, "_apply_primary_worker") as apply_worker_mock,
            patch.object(service, "_save_batch_row") as save_row_mock,
        ):
            errors, participant_updated = service._apply_row(
                uploaded_row,
                batch=MagicMock(),
                upload_date=date.today(),
                uploaded_at=datetime.now(),
            )

        self.assertEqual(errors, [])
        self.assertFalse(participant_updated)
        apply_worker_mock.assert_not_called()
        self.assertEqual(
            save_row_mock.call_args.kwargs["status"],
            HouseholdValidationBatchRow.Status.SKIPPED,
        )

    def test_only_valid_verified_rows_enable_household_participant_updates(self):
        verified_row = self._uploaded_row(2, "member-1", True)
        unverified_row = UploadedValidationRow(
            row_number=3,
            values={
                "group_uuid": "group-2",
                "member_uuid": "member-2",
                "form_number": "FORM-002",
            },
            verified=None,
            primary_worker=True,
            validation_date=None,
            project_name=None,
            project_id=None,
            notes=None,
        )
        invalid_verified_row = UploadedValidationRow(
            row_number=4,
            values={
                "group_uuid": "group-3",
                "member_uuid": "member-3",
                "form_number": "FORM-003",
            },
            verified=True,
            primary_worker=True,
            validation_date=None,
            project_name=None,
            project_id=None,
            notes=None,
        )
        service = HouseholdValidationUploadService()

        with patch.object(
            service,
            "_resolve_row",
            side_effect=[
                ([], MagicMock(), MagicMock(), None),
                (["Row 4: member was not found in group"], None, None, None),
            ],
        ):
            accepted = service._accepted_validation_group_keys(
                [verified_row, unverified_row, invalid_verified_row]
            )

        self.assertEqual(accepted, {"group-1"})

    @patch("household_validation.services.parse_validation_workbook")
    def test_upload_counts_only_actual_participant_changes(self, parse_workbook_mock):
        rows = [
            self._uploaded_row(2, "member-1", True),
            self._uploaded_row(3, "member-2", False),
            self._uploaded_row(4, "member-3", None),
        ]
        parse_workbook_mock.return_value = SimpleNamespace(
            rows=rows,
            rows_read=3,
            errors=[],
        )
        batch = MagicMock(id="batch-1")
        service = HouseholdValidationUploadService()

        with (
            patch.object(service, "_primary_worker_rejections", return_value={}),
            patch.object(service, "_get_or_create_batch", return_value=batch),
            patch.object(
                service,
                "_apply_row",
                side_effect=[([], False), ([], True), ([], False)],
            ),
            patch(
                "household_validation.services.transaction.atomic",
                return_value=nullcontext(),
            ),
        ):
            result = service.upload(b"workbook")

        self.assertEqual(result["participant_updates"], 1)
        self.assertEqual(result["households_verified"], 1)
        self.assertEqual(result["households_not_verified"], 0)

    @patch("household_validation.services.Group.objects.filter")
    def test_group_lookup_prefetches_members_and_individuals(self, filter_mock):
        group = SimpleNamespace(id="group-1")
        filtered_queryset = MagicMock()
        location_queryset = MagicMock()
        prefetched_queryset = MagicMock()
        filter_mock.return_value = filtered_queryset
        filtered_queryset.select_related.return_value = location_queryset
        location_queryset.prefetch_related.return_value = prefetched_queryset
        prefetched_queryset.first.return_value = group

        service = HouseholdValidationUploadService()
        result = service._group("group-1")
        cached_result = service._group("group-1")

        self.assertIs(result, group)
        self.assertIs(cached_result, group)
        filter_mock.assert_called_once_with(id="group-1")
        filtered_queryset.select_related.assert_called_once_with("location")
        location_queryset.prefetch_related.assert_called_once_with(
            "groupindividuals__individual"
        )

    def test_group_individual_lookup_reuses_prefetched_members(self):
        member = SimpleNamespace(
            individual_id="individual-1",
            is_deleted=False,
        )
        group = SimpleNamespace(
            _prefetched_objects_cache={"groupindividuals": [member]},
        )
        service = HouseholdValidationUploadService()

        with patch(
            "household_validation.services.GroupIndividual.objects.filter"
        ) as filter_mock:
            result = service._group_individual("individual-1", group)

        self.assertIs(result, member)
        filter_mock.assert_not_called()

    @patch("household_validation.services.GroupIndividual.objects.filter")
    def test_group_individual_lookup_uses_exported_individual_id(
        self,
        filter_mock,
    ):
        group = SimpleNamespace(id="group-1")
        group_individual = SimpleNamespace(
            id="membership-1",
            individual_id="individual-1",
            group=group,
        )
        queryset = MagicMock()
        queryset.select_related.return_value.first.return_value = group_individual
        filter_mock.return_value = queryset

        result = HouseholdValidationUploadService()._group_individual(
            "individual-1",
            group=group,
        )

        self.assertIs(result, group_individual)
        filter_mock.assert_called_once_with(
            individual_id="individual-1",
            group=group,
            is_deleted=False,
        )
        queryset.select_related.assert_called_once_with("individual")

    def test_json_safe_converts_nested_dates_to_iso_strings(self):
        raw_row = {
            "member_dob": date(1963, 10, 21),
            "validation": {
                "validation_date": date(2026, 7, 29),
            },
        }

        serialized = _json_safe(raw_row)

        self.assertEqual(serialized["member_dob"], "1963-10-21")
        self.assertEqual(
            serialized["validation"]["validation_date"],
            "2026-07-29",
        )

    def test_build_validation_json_ext_persists_not_verified_status(self):
        uploaded_row = UploadedValidationRow(
            row_number=2,
            values={},
            verified=False,
            primary_worker=None,
            validation_date=date(2026, 7, 3),
            project_name="Road Works",
            project_id="project-1",
            notes="Not available",
        )

        json_ext = build_validation_json_ext(
            uploaded_row=uploaded_row,
            project=None,
            upload_date=date(2026, 7, 4),
            uploaded_at=SimpleNamespace(isoformat=lambda: "2026-07-04T09:00:00+00:00"),
            user_id="user-1",
        )

        self.assertEqual(json_ext["validation_status"], VALIDATION_STATUS_NOT_VERIFIED)
        self.assertEqual(json_ext["last_verified_date"], "2026-07-03")
        self.assertEqual(json_ext["validation_project_selection_type"], PROJECT_SELECTION_TYPE_INTENT)

    def test_member_structural_errors_detect_protected_field_tampering(self):
        group = SimpleNamespace(
            json_ext={
                "form_number": "FORM-001",
                "household_wealth_quintile": "Poorest",
            },
        )
        group_individual = SimpleNamespace(
            role="HEAD",
            recipient_type="PRIMARY",
            group=group,
            individual=SimpleNamespace(
                first_name="Ada",
                last_name="Worker",
                dob=date(1990, 1, 1),
                json_ext={
                    "gender": "Female",
                    "fit_for_work": True,
                    "national_id": "NAT-001",
                },
            ),
        )
        uploaded_row = UploadedValidationRow(
            row_number=2,
            values={
                "form_number": "FORM-999",
                "member_name": "Grace Worker",
                "national_id": "NAT-999",
                "member_gender": "Male",
                "member_dob": "1991-01-01",
                "member_age": "18",
                "fit_for_work": "NO",
                "relationship": "SPOUSE",
                "head": "NO",
                "household_wealth_quintile": "Richest",
            },
            verified=None,
            primary_worker=None,
            validation_date=None,
            project_name=None,
            project_id=None,
            notes=None,
        )

        errors = member_structural_errors(uploaded_row, group_individual)

        self.assertIn("Row 2: form_number does not match the household member", errors)
        self.assertIn("Row 2: member_name does not match the household member", errors)
        self.assertNotIn("Row 2: national_id does not match the household member", errors)
        self.assertIn("Row 2: member_gender does not match the household member", errors)
        self.assertIn("Row 2: fit_for_work does not match the household member", errors)
        self.assertIn("Row 2: relationship does not match the household member", errors)
        self.assertIn(
            "Row 2: household_wealth_quintile does not match the household member",
            errors,
        )

    def test_structural_errors_accept_integer_like_excel_ids(self):
        group = SimpleNamespace(
            json_ext={
                "form_number": "123456",
                "household_wealth_quintile": "Poorest",
            },
        )
        group_individual = SimpleNamespace(
            role="HEAD",
            recipient_type=None,
            group=group,
            individual=SimpleNamespace(
                first_name="Ada",
                last_name="Worker",
                dob=None,
                json_ext={
                    "fit_for_work": True,
                    "national_id": "123456",
                },
            ),
        )
        uploaded_row = UploadedValidationRow(
            row_number=2,
            values={
                "form_number": 123456.0,
                "national_id": 123456.0,
                "relationship": "HEAD",
                "household_wealth_quintile": "Poorest",
            },
            verified=None,
            primary_worker=None,
            validation_date=None,
            project_name=None,
            project_id=None,
            notes=None,
        )

        errors = member_structural_errors(uploaded_row, group_individual)

        self.assertNotIn(
            "Row 2: form_number does not match the household member",
            errors,
        )
        self.assertNotIn(
            "Row 2: national_id does not match the household member",
            errors,
        )

    def test_form_number_maps_to_group_without_using_group_code(self):
        group = SimpleNamespace(
            code="INTERNAL-GROUP-CODE",
            json_ext={},
        )
        group_individual = SimpleNamespace(
            role="HEAD",
            recipient_type=None,
            group=group,
            individual=SimpleNamespace(
                first_name="Ada",
                last_name="Worker",
                dob=None,
                json_ext={
                    "fit_for_work": True,
                    "form_number": "FORM-001",
                    "national_id": "NAT-001",
                },
            ),
        )
        group.groupindividuals = _FakeRelatedManager([group_individual])
        uploaded_row = UploadedValidationRow(
            row_number=2,
            values={
                "form_number": "FORM-001",
                "national_id": "NAT-001",
                "relationship": "HEAD",
            },
            verified=None,
            primary_worker=None,
            validation_date=None,
            project_name=None,
            project_id=None,
            notes=None,
        )

        errors = member_structural_errors(
            uploaded_row,
            group_individual,
            group=group,
        )

        self.assertNotIn(
            "Row 2: form_number does not match the household member",
            errors,
        )
        self.assertEqual(group.code, "INTERNAL-GROUP-CODE")

    def test_upload_wealth_validation_uses_head_before_first_member(self):
        other_member = self._wealth_member(
            "other",
            role="SPOUSE",
            fit_for_work=True,
            wealth_quintile="Richest",
        )
        head = self._wealth_member(
            "head",
            role="HEAD",
            fit_for_work=True,
            wealth_quintile="Poorer",
        )
        group = SimpleNamespace(
            json_ext={},
            groupindividuals=_FakeRelatedManager([other_member, head]),
        )
        other_member.group = group
        uploaded_row = UploadedValidationRow(
            row_number=2,
            values={
                "relationship": "SPOUSE",
                "household_wealth_quintile": "Poorer",
            },
            verified=None,
            primary_worker=None,
            validation_date=None,
            project_name=None,
            project_id=None,
            notes=None,
        )

        errors = member_structural_errors(uploaded_row, other_member)

        self.assertNotIn(
            "Row 2: household_wealth_quintile does not match "
            "the household member",
            errors,
        )

    def test_wealth_resolver_ignores_members_who_are_not_fit_for_work(self):
        ineligible_member = self._wealth_member(
            "ineligible",
            role="SPOUSE",
            fit_for_work=False,
            wealth_quintile="Richest",
        )
        eligible_member = self._wealth_member(
            "eligible",
            role="SON",
            fit_for_work=True,
            wealth_quintile="Middle",
        )
        group = SimpleNamespace(
            json_ext={},
            groupindividuals=_FakeRelatedManager(
                [ineligible_member, eligible_member]
            ),
        )

        self.assertEqual(get_household_wealth_quintile(group), "Middle")

    def test_pmt_score_resolver_ignores_members_who_are_not_fit_for_work(self):
        ineligible_member = self._wealth_member(
            "ineligible",
            role="SPOUSE",
            fit_for_work=False,
            wealth_quintile="Richest",
            pmt_score=9.9,
        )
        eligible_member = self._wealth_member(
            "eligible",
            role="SON",
            fit_for_work=True,
            wealth_quintile="Middle",
            pmt_score=-1.5,
        )
        group = SimpleNamespace(
            json_ext={},
            groupindividuals=_FakeRelatedManager(
                [ineligible_member, eligible_member]
            ),
        )

        self.assertEqual(get_household_pmt_score(group), -1.5)

    def test_export_selection_uses_shared_wealth_precedence(self):
        other_member = self._wealth_member(
            "other",
            role="SPOUSE",
            fit_for_work=True,
            wealth_quintile="Richest",
            pmt_score=9.9,
        )
        head = self._wealth_member(
            "head",
            role="HEAD",
            fit_for_work=True,
            wealth_quintile="Poorest",
            pmt_score=-2.1,
        )
        group = SimpleNamespace(
            id="group-1",
            code="HH-001",
            json_ext={},
            groupindividuals=_FakeRelatedManager([other_member, head]),
        )

        household = EligibleHouseholdSelectionService()._build_household(group)

        self.assertEqual(household.wealth_quintile, "Poorest")
        self.assertEqual(household.pmt_score, -2.1)

    def _wealth_member(self, member_id, role, fit_for_work, wealth_quintile, pmt_score=None):
        json_ext = {
            "fit_for_work": fit_for_work,
            "household_wealth_quintile": wealth_quintile,
        }
        if pmt_score is not None:
            json_ext["household_pmt_score"] = pmt_score
        individual = SimpleNamespace(
            id=f"individual-{member_id}",
            dob=date(1990, 1, 1),
            json_ext=json_ext,
        )
        return SimpleNamespace(
            id=f"membership-{member_id}",
            individual_id=individual.id,
            individual=individual,
            role=role,
            recipient_type=None,
        )

    def test_build_validation_error_report_writes_error_rows_csv(self):
        batch = SimpleNamespace(
            id="batch-1",
            rows=_FakeRows(
                [
                    SimpleNamespace(
                        row_number=2,
                        status="ERROR",
                        raw_row={
                            "form_number": "FORM-001",
                            "group_uuid": "group-1",
                            "member_uuid": "member-1",
                        },
                        error_message="form_number does not match",
                    )
                ]
            ),
        )

        report = build_validation_error_report_csv(batch.id, batch.rows.order_by("row_number"))

        self.assertIn("batch_id,row_number,status,form_number,group_uuid,member_uuid,error_message", report)
        self.assertIn("batch-1,2,ERROR,FORM-001,group-1,member-1,form_number does not match", report)

    def test_error_report_uses_legacy_group_code_as_form_number(self):
        row = SimpleNamespace(
            row_number=2,
            status="ERROR",
            raw_row={
                "group_code": "LEGACY-001",
                "group_uuid": "group-1",
                "member_uuid": "member-1",
            },
            error_message="legacy upload error",
        )

        report = build_validation_error_report_csv("batch-1", [row])

        self.assertIn(
            "batch-1,2,ERROR,LEGACY-001,group-1,member-1,legacy upload error",
            report,
        )


class LocalDateTests(TestCase):
    """openIMIS runs with USE_TZ = False, so timezone.now() is naive.

    Passing a naive datetime to timezone.localdate()/localtime() raises
    ValueError("localtime() cannot be applied to a naive datetime"), which
    previously aborted every non-dry-run validation-list upload.
    """

    def test_naive_datetime_returns_its_date(self):
        naive = datetime(2026, 7, 29, 12, 27)
        self.assertIsNone(naive.tzinfo)
        self.assertEqual(_local_date(naive), date(2026, 7, 29))

    def test_aware_datetime_is_converted_to_local_date(self):
        aware = datetime(2026, 7, 29, 12, 27, tzinfo=datetime_timezone.utc)
        self.assertEqual(_local_date(aware), timezone.localtime(aware).date())


class _FakeRows:
    def __init__(self, rows):
        self.rows = rows

    def filter(self, **kwargs):
        return self

    def order_by(self, *fields):
        return self.rows
